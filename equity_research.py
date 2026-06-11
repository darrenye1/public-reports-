#!/usr/bin/env python3
"""
Automated Equity Research Report Generator (single-file)

Fetches data from Yahoo Finance, performs financial/industry/competitor analysis,
runs DCF and comparable valuation, and outputs institutional-style PDF reports.

Usage:
    python equity_research.py AAPL
    python equity_research.py --batch-top20
    python equity_research.py --batch-top20 --weekly
"""

from __future__ import annotations

import argparse
import glob
import io
import json
import os
import statistics
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
import yfinance as yf
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

DEFAULT_WACC = 0.10
DEFAULT_TERMINAL_GROWTH = 0.025
DEFAULT_PROJECTION_YEARS = 5
DEFAULT_FCF_GROWTH = 0.08
DEFAULT_OUTPUT_DIR = "reports"
TOP_US_PDF_REPORT_COUNT = 5
TOP_US_SUMMARY_COUNT = 20
SUMMARY_REPORT_NAME = "Top20_Summary.pdf"

# Always generate full PDF reports for these tickers (in addition to top-N by market cap).
# Blackstone Group = BX (not BSX, which is Boston Scientific).
EXTRA_PDF_TICKERS: list[str] = ["TSLA", "BX"]

# Broad US universe; ranked by live market cap.
MEGA_CAP_UNIVERSE: list[str] = [
    "NVDA", "AAPL", "MSFT", "GOOGL", "GOOG", "AMZN", "META", "BRK-B", "AVGO", "TSLA",
    "WMT", "JPM", "LLY", "V", "ORCL", "XOM", "MA", "UNH", "COST", "HD", "PG", "NFLX",
    "BAC", "ABBV", "CRM", "KO", "AMD", "PEP", "CVX", "LIN", "TMO", "ADBE", "DIS", "CSCO",
    "MCD", "ABT", "ACN", "INTU", "WFC", "MRK", "GE", "PM", "TXN", "QCOM", "IBM", "CAT",
    "AMAT", "NOW", "UBER", "ISRG", "GS", "AXP", "BKNG", "SPGI", "MS", "RTX", "PFE", "LOW",
    "SYK", "VZ", "T", "CMCSA", "HON", "NEE", "UNP", "DHR", "AMGN", "PGR", "BLK", "BX", "ELV",
    "DE", "MDT", "ADI", "CI", "GILD", "ADP", "C", "TJX", "SO", "ZTS", "MO", "MMC", "CB",
    "BSX", "PLD", "REGN", "SHW", "MU", "ICE", "DUK", "CL", "BMY", "NKE", "SBUX",
]

# Outlier thresholds when cross-checking valuation methods.
OUTLIER_VS_MEDIAN_LOW = 0.52
OUTLIER_VS_MEDIAN_HIGH = 1.72
OUTLIER_VS_CURRENT_LOW = 0.42
OUTLIER_VS_CURRENT_HIGH = 2.0

DATA_STATE_FILENAME = ".data_state.json"
BATCH_TICKER_DELAY_SEC = 2.0
UNIVERSE_FETCH_DELAY_SEC = 0.35
YFINANCE_FETCH_RETRIES = 6
YFINANCE_RETRY_BASE_SEC = 4.0
STALE_MAX_AGE_DAYS = 7

SECTOR_PEERS: dict[str, list[str]] = {
    "Technology": ["MSFT", "GOOGL", "META", "NVDA", "ORCL", "CRM", "ADBE"],
    "Consumer Cyclical": ["AMZN", "TSLA", "HD", "NKE", "SBUX", "MCD"],
    "Auto Manufacturers": ["F", "GM", "RIVN", "TM", "HMC", "STLA", "LI", "NIO"],
    "Healthcare": ["JNJ", "UNH", "PFE", "ABBV", "MRK", "LLY"],
    "Financial Services": ["JPM", "BAC", "WFC", "GS", "MS", "BLK"],
    "Communication Services": ["GOOGL", "META", "DIS", "NFLX", "CMCSA"],
    "Industrials": ["CAT", "GE", "HON", "UPS", "BA", "DE"],
    "Energy": ["XOM", "CVX", "COP", "SLB", "EOG"],
    "Consumer Defensive": ["PG", "KO", "PEP", "WMT", "COST"],
}

NAVY = colors.HexColor("#1a365d")
ACCENT = colors.HexColor("#2b6cb0")
LIGHT_BG = colors.HexColor("#f7fafc")
BORDER = colors.HexColor("#cbd5e0")
GREEN = colors.HexColor("#276749")
RED = colors.HexColor("#c53030")


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------


@dataclass
class CompanyData:
    ticker: str
    info: dict[str, Any] = field(default_factory=dict)
    financials: pd.DataFrame | None = None
    balance_sheet: pd.DataFrame | None = None
    cashflow: pd.DataFrame | None = None
    history: pd.DataFrame | None = None
    recommendations: pd.DataFrame | None = None


@dataclass
class FinancialSummary:
    company_name: str
    ticker: str
    sector: str
    industry: str
    market_cap: float | None
    current_price: float | None
    revenue: float | None
    revenue_growth: float | None
    gross_margin: float | None
    operating_margin: float | None
    profit_margin: float | None
    eps: float | None
    pe_ratio: float | None
    roe: float | None
    debt_to_equity: float | None
    free_cash_flow: float | None
    ebitda: float | None
    ev_ebitda: float | None
    beta: float | None
    dividend_yield: float | None
    fifty_two_week_range: str
    analyst_target: float | None
    recommendation_key: str


@dataclass
class IndustryResearch:
    sector: str
    industry: str
    sector_pe: float | None
    industry_pe: float | None
    market_cap_category: str
    business_summary: str
    key_trends: list[str]
    risks: list[str]


@dataclass
class PeerMetrics:
    ticker: str
    name: str
    market_cap: float | None
    pe_ratio: float | None
    ev_ebitda: float | None
    profit_margin: float | None
    revenue_growth: float | None
    roe: float | None


@dataclass
class CompetitorAnalysis:
    target_ticker: str
    peers: list[PeerMetrics]
    peer_median_pe: float | None
    peer_median_ev_ebitda: float | None
    peer_median_margin: float | None
    relative_position: str


@dataclass
class DCFResult:
    fcf_base: float | None
    wacc: float  # cost of equity (Ke) in Simon FCFE model
    terminal_growth: float
    projection_years: int
    fcf_growth: float
    enterprise_value: float | None
    equity_value: float | None
    implied_price: float | None
    current_price: float | None
    upside_pct: float | None
    assumptions_note: str
    projected_fcfe: list[float] = field(default_factory=list)
    pv_by_year: list[float] = field(default_factory=list)
    pv_explicit: float | None = None
    pv_terminal: float | None = None
    terminal_fcfe: float | None = None
    equity_value_raw: float | None = None
    shares_used: float | None = None


@dataclass
class CompsResult:
    peer_median_pe: float | None
    peer_median_ev_ebitda: float | None
    peer_median_ps: float | None
    implied_price_pe: float | None
    implied_price_ev_ebitda: float | None
    implied_price_ps: float | None
    blended_fair_value: float | None
    current_price: float | None
    upside_pct: float | None


@dataclass
class ValuationSummary:
    dcf: DCFResult
    comps: CompsResult
    recommendation: str
    target_price: float | None
    dcf_target: float | None = None
    comps_target: float | None = None
    analyst_target: float | None = None
    dcf_reliable: bool = True
    dcf_exclusion_reason: str = ""
    comps_reliable: bool = True
    analyst_reliable: bool = True
    excluded_methods: dict[str, str] = field(default_factory=dict)
    blend_weights: dict[str, float] = field(default_factory=dict)


@dataclass
class ReportRow:
    ticker: str
    company_name: str
    sector: str
    current_price: float | None
    market_cap: float | None
    pe_ratio: float | None
    revenue_growth: float | None
    operating_margin: float | None
    roe: float | None
    dcf_fair_value: float | None
    comps_fair_value: float | None
    target_price: float | None
    upside_pct: float | None
    rating: str
    financial_health: str
    output_path: str


# ---------------------------------------------------------------------------
# Data fetching helpers
# ---------------------------------------------------------------------------


def _safe_series_value(df: pd.DataFrame | None, row_label: str, col_index: int = 0) -> float | None:
    if df is None or df.empty:
        return None
    if row_label not in df.index:
        matches = [idx for idx in df.index if row_label.lower() in str(idx).lower()]
        if not matches:
            return None
        row_label = matches[0]
    try:
        val = df.loc[row_label].iloc[col_index]
        if pd.isna(val):
            return None
        return float(val)
    except (IndexError, KeyError, TypeError, ValueError):
        return None


def get_info_value(info: dict, *keys: str, default: Any = None) -> Any:
    for key in keys:
        val = info.get(key)
        if val is not None:
            return val
    return default


def _yfinance_retryable(exc: BaseException) -> bool:
    name = type(exc).__name__
    if name in ("YFRateLimitError", "HTTPError", "ConnectionError", "TimeoutError"):
        return True
    msg = str(exc).lower()
    return "rate limit" in msg or "too many requests" in msg or "429" in msg


def fetch_company(ticker: str, history_period: str = "1y") -> CompanyData:
    symbol = ticker.upper().strip()
    last_err: BaseException | None = None
    for attempt in range(YFINANCE_FETCH_RETRIES):
        try:
            stock = yf.Ticker(symbol)
            info = stock.info or {}
            if not info or info.get("regularMarketPrice") is None and info.get("currentPrice") is None:
                hist = stock.history(period="5d")
                if hist.empty:
                    raise ValueError(
                        f"Could not fetch data for ticker '{symbol}'. Check symbol and try again."
                    )
            return CompanyData(
                ticker=symbol,
                info=info,
                financials=stock.financials,
                balance_sheet=stock.balance_sheet,
                cashflow=stock.cashflow,
                history=stock.history(period=history_period),
                recommendations=getattr(stock, "recommendations", None),
            )
        except ValueError:
            raise
        except Exception as e:
            last_err = e
            if attempt < YFINANCE_FETCH_RETRIES - 1 and _yfinance_retryable(e):
                wait = YFINANCE_RETRY_BASE_SEC * (attempt + 1)
                print(
                    f"  Yahoo Finance rate limit for {symbol}; retry in {wait:.0f}s...",
                    file=sys.stderr,
                )
                time.sleep(wait)
                continue
            raise
    if last_err:
        raise last_err
    raise ValueError(f"Could not fetch data for ticker '{symbol}'.")


def get_financial_period_key(data: CompanyData) -> str | None:
    """Latest reported period from statements (e.g. FY2024 or Q3 FY2024)."""
    for df in (data.financials, data.cashflow, data.balance_sheet):
        if df is None or df.empty:
            continue
        col = df.columns[0]
        try:
            if hasattr(col, "year"):
                q = getattr(col, "quarter", None)
                if q:
                    return f"Q{int(q)} FY{col.year}"
                return f"FY{col.year}"
        except (TypeError, ValueError):
            pass
        return str(col)[:12]
    mrq = get_info_value(data.info, "mostRecentQuarter")
    if mrq:
        try:
            ts = float(mrq)
            if ts > 1e9:
                return datetime.utcfromtimestamp(ts).strftime("%Y-%m-%d")
        except (TypeError, ValueError):
            return str(mrq)[:12]
    return None


def data_state_path(output_dir: str) -> str:
    return os.path.join(output_dir, DATA_STATE_FILENAME)


def load_data_state(output_dir: str) -> dict[str, Any]:
    path = data_state_path(output_dir)
    if not os.path.isfile(path):
        return {"tickers": {}, "summary_last_run": None}
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError):
        return {"tickers": {}, "summary_last_run": None}


def save_data_state(output_dir: str, state: dict[str, Any]) -> None:
    os.makedirs(output_dir, exist_ok=True)
    path = data_state_path(output_dir)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)


def _report_row_to_dict(row: ReportRow) -> dict[str, Any]:
    return {
        "ticker": row.ticker,
        "company_name": row.company_name,
        "sector": row.sector,
        "current_price": row.current_price,
        "market_cap": row.market_cap,
        "pe_ratio": row.pe_ratio,
        "revenue_growth": row.revenue_growth,
        "operating_margin": row.operating_margin,
        "roe": row.roe,
        "dcf_fair_value": row.dcf_fair_value,
        "comps_fair_value": row.comps_fair_value,
        "target_price": row.target_price,
        "upside_pct": row.upside_pct,
        "rating": row.rating,
        "financial_health": row.financial_health,
        "output_path": row.output_path,
    }


def _report_row_from_dict(d: dict[str, Any]) -> ReportRow:
    return ReportRow(
        ticker=d["ticker"],
        company_name=d.get("company_name", d["ticker"]),
        sector=d.get("sector", ""),
        current_price=d.get("current_price"),
        market_cap=d.get("market_cap"),
        pe_ratio=d.get("pe_ratio"),
        revenue_growth=d.get("revenue_growth"),
        operating_margin=d.get("operating_margin"),
        roe=d.get("roe"),
        dcf_fair_value=d.get("dcf_fair_value"),
        comps_fair_value=d.get("comps_fair_value"),
        target_price=d.get("target_price"),
        upside_pct=d.get("upside_pct"),
        rating=d.get("rating", "HOLD"),
        financial_health=d.get("financial_health", ""),
        output_path=d.get("output_path", ""),
    )


def confirm_ticker_refresh(
    ticker: str,
    state: dict[str, Any],
    data: CompanyData,
    force: bool,
    max_age_days: int = STALE_MAX_AGE_DAYS,
) -> tuple[bool, str]:
    """After fetching data, decide if report should be regenerated."""
    symbol = ticker.upper()
    if force:
        return True, "forced refresh"
    entry = state.get("tickers", {}).get(symbol)
    if not entry:
        return True, "no prior run"
    period = get_financial_period_key(data)
    if period and period != entry.get("financial_period"):
        return True, f"new financial period ({period})"
    last_run = entry.get("last_run_utc")
    if last_run:
        try:
            last_dt = datetime.fromisoformat(last_run.replace("Z", "+00:00"))
            if last_dt.tzinfo is None:
                last_dt = last_dt.replace(tzinfo=timezone.utc)
            age_days = (datetime.now(timezone.utc) - last_dt).days
            if age_days >= max_age_days:
                return True, f"older than {max_age_days} days ({age_days}d)"
        except ValueError:
            return True, "invalid cached timestamp"
    else:
        return True, "missing last run timestamp"
    return False, "up to date"


def _cached_report_row(state: dict[str, Any], ticker: str) -> ReportRow | None:
    entry = state.get("tickers", {}).get(ticker.upper(), {})
    raw = entry.get("summary_row")
    if raw:
        return _report_row_from_dict(raw)
    return None


def record_ticker_state(
    state: dict[str, Any],
    ticker: str,
    data: CompanyData,
    row: ReportRow | None,
) -> None:
    symbol = ticker.upper()
    tickers = state.setdefault("tickers", {})
    tickers[symbol] = {
        "last_run_utc": datetime.now(timezone.utc).isoformat(),
        "financial_period": get_financial_period_key(data),
        "data_as_of": datetime.now().strftime("%Y-%m-%d"),
        "summary_row": _report_row_to_dict(row) if row else tickers.get(symbol, {}).get("summary_row"),
    }


def get_free_cash_flow(cashflow: pd.DataFrame | None) -> float | None:
    fcf = _safe_series_value(cashflow, "Free Cash Flow")
    if fcf is not None:
        return fcf
    ocf = _safe_series_value(cashflow, "Operating Cash Flow")
    capex = _safe_series_value(cashflow, "Capital Expenditure")
    if ocf is not None and capex is not None:
        return ocf + capex if capex < 0 else ocf - capex
    return None


def get_net_debt(balance_sheet: pd.DataFrame | None) -> float | None:
    if balance_sheet is None or balance_sheet.empty:
        return None
    total_debt = _safe_series_value(balance_sheet, "Total Debt")
    if total_debt is None:
        long_debt = _safe_series_value(balance_sheet, "Long Term Debt") or 0
        short_debt = _safe_series_value(balance_sheet, "Current Debt") or 0
        total_debt = long_debt + short_debt if (long_debt or short_debt) else None
    cash = _safe_series_value(balance_sheet, "Cash And Cash Equivalents")
    if cash is None:
        cash = _safe_series_value(balance_sheet, "Cash Cash Equivalents And Short Term Investments") or 0
    if total_debt is None:
        return None
    return total_debt - (cash or 0)


def get_revenue(financials: pd.DataFrame | None, info: dict) -> float | None:
    rev = _safe_series_value(financials, "Total Revenue")
    if rev is None:
        rev = get_info_value(info, "totalRevenue")
    return float(rev) if rev is not None else None


def get_ebitda(financials: pd.DataFrame | None, info: dict) -> float | None:
    ebitda = _safe_series_value(financials, "EBITDA")
    if ebitda is None:
        ebitda = get_info_value(info, "ebitda")
    return float(ebitda) if ebitda is not None else None


def get_shares_outstanding(info: dict) -> float | None:
    shares = get_info_value(info, "sharesOutstanding", "impliedSharesOutstanding")
    return float(shares) if shares else None


def get_current_price(info: dict, history: pd.DataFrame | None = None) -> float | None:
    price = get_info_value(info, "currentPrice", "regularMarketPrice", "previousClose")
    if price is not None:
        return float(price)
    if history is not None and not history.empty:
        return float(history["Close"].iloc[-1])
    return None


def discover_sector_peers(ticker: str, sector: str | None, industry: str | None, max_peers: int = 5) -> list[str]:
    symbol = ticker.upper()
    candidates: list[str] = []
    if industry and industry in SECTOR_PEERS:
        candidates = [p for p in SECTOR_PEERS[industry] if p != symbol]
    elif industry:
        for key, peers in SECTOR_PEERS.items():
            if key.lower() in industry.lower() or industry.lower() in key.lower():
                candidates = [p for p in peers if p != symbol]
                break
    if not candidates and sector and sector in SECTOR_PEERS:
        candidates = [p for p in SECTOR_PEERS[sector] if p != symbol]
    elif not candidates and sector:
        for key, peers in SECTOR_PEERS.items():
            if sector.lower() in key.lower() or key.lower() in sector.lower():
                candidates = [p for p in peers if p != symbol]
                break
    if not candidates:
        candidates = ["SPY"]
    return candidates[:max_peers]


def format_large_number(value: float | None, prefix: str = "$") -> str:
    if value is None:
        return "N/A"
    abs_val = abs(value)
    if abs_val >= 1e12:
        return f"{prefix}{value / 1e12:.2f}T"
    if abs_val >= 1e9:
        return f"{prefix}{value / 1e9:.2f}B"
    if abs_val >= 1e6:
        return f"{prefix}{value / 1e6:.2f}M"
    return f"{prefix}{value:,.2f}"


def _pct(value: float | None) -> float | None:
    if value is None:
        return None
    return value * 100 if abs(value) <= 1 else value


# ---------------------------------------------------------------------------
# Analysis
# ---------------------------------------------------------------------------


def build_financial_summary(data: CompanyData) -> FinancialSummary:
    info = data.info
    fin = data.financials
    cf = data.cashflow

    rev_t = get_revenue(fin, info)
    rev_t1 = _safe_series_value(fin, "Total Revenue", 1) if fin is not None else None
    rev_growth = None
    if rev_t and rev_t1 and rev_t1 != 0:
        rev_growth = (rev_t - rev_t1) / abs(rev_t1) * 100

    low = get_info_value(info, "fiftyTwoWeekLow")
    high = get_info_value(info, "fiftyTwoWeekHigh")
    range_str = f"${low:.2f} – ${high:.2f}" if low and high else "N/A"

    fcf = None
    if cf is not None and not cf.empty:
        fcf = _safe_series_value(cf, "Free Cash Flow")
        if fcf is None:
            ocf = _safe_series_value(cf, "Operating Cash Flow")
            capex = _safe_series_value(cf, "Capital Expenditure")
            if ocf is not None and capex is not None:
                fcf = ocf + capex if capex < 0 else ocf - capex

    return FinancialSummary(
        company_name=get_info_value(info, "longName", "shortName", default=data.ticker) or data.ticker,
        ticker=data.ticker,
        sector=get_info_value(info, "sector", default="N/A") or "N/A",
        industry=get_info_value(info, "industry", default="N/A") or "N/A",
        market_cap=get_info_value(info, "marketCap"),
        current_price=get_current_price(info, data.history),
        revenue=rev_t,
        revenue_growth=rev_growth,
        gross_margin=_pct(get_info_value(info, "grossMargins")),
        operating_margin=_pct(get_info_value(info, "operatingMargins")),
        profit_margin=_pct(get_info_value(info, "profitMargins")),
        eps=get_info_value(info, "trailingEps"),
        pe_ratio=get_info_value(info, "trailingPE"),
        roe=_pct(get_info_value(info, "returnOnEquity")),
        debt_to_equity=get_info_value(info, "debtToEquity"),
        free_cash_flow=fcf,
        ebitda=get_ebitda(fin, info),
        ev_ebitda=get_info_value(info, "enterpriseToEbitda"),
        beta=get_info_value(info, "beta"),
        dividend_yield=_pct(get_info_value(info, "dividendYield")),
        fifty_two_week_range=range_str,
        analyst_target=get_info_value(info, "targetMeanPrice"),
        recommendation_key=get_info_value(info, "recommendationKey", default="N/A") or "N/A",
    )


def _truncate(text: str, max_len: int = 400) -> str:
    if not text:
        return "No business summary available from data provider."
    text = " ".join(text.split())
    if len(text) <= max_len:
        return text
    return text[: max_len - 3].rsplit(" ", 1)[0] + "..."


def build_industry_research(data: CompanyData, summary: FinancialSummary) -> IndustryResearch:
    info = data.info
    sector = summary.sector
    industry = summary.industry

    mcap = summary.market_cap
    if mcap and mcap >= 200e9:
        cap_cat = "Large Cap (>$200B)"
    elif mcap and mcap >= 10e9:
        cap_cat = "Mid Cap ($10B–$200B)"
    elif mcap and mcap >= 2e9:
        cap_cat = "Small Cap ($2B–$10B)"
    else:
        cap_cat = "Micro/Small Cap (<$2B)"

    business = get_info_value(info, "longBusinessSummary", default="") or ""
    trends: list[str] = []
    if summary.revenue_growth is not None:
        if summary.revenue_growth > 15:
            trends.append(f"Strong top-line growth ({summary.revenue_growth:.1f}% YoY)")
        elif summary.revenue_growth > 0:
            trends.append(f"Moderate revenue expansion ({summary.revenue_growth:.1f}% YoY)")
        else:
            trends.append(f"Revenue contraction ({summary.revenue_growth:.1f}% YoY)")

    if summary.operating_margin is not None:
        if summary.operating_margin > 25:
            trends.append(f"High operating leverage ({summary.operating_margin:.1f}% margin)")
        elif summary.operating_margin > 10:
            trends.append(f"Healthy operating margins ({summary.operating_margin:.1f}%)")

    if summary.beta is not None:
        if summary.beta > 1.2:
            trends.append(f"Above-market volatility (β={summary.beta:.2f})")
        elif summary.beta < 0.8:
            trends.append(f"Defensive profile (β={summary.beta:.2f})")

    trends.append(f"Sector: {sector} | Industry: {industry}")

    risks: list[str] = []
    if summary.debt_to_equity and summary.debt_to_equity > 100:
        risks.append(f"Elevated leverage (D/E: {summary.debt_to_equity:.0f}%)")
    if summary.pe_ratio and summary.pe_ratio > 40:
        risks.append(f"Premium valuation (P/E: {summary.pe_ratio:.1f}x)")
    if summary.beta and summary.beta > 1.5:
        risks.append("High market sensitivity in downturns")
    if not risks:
        risks.append("Standard sector cyclicality and macro exposure")

    return IndustryResearch(
        sector=sector,
        industry=industry,
        sector_pe=get_info_value(info, "trailingPE"),
        industry_pe=None,
        market_cap_category=cap_cat,
        business_summary=_truncate(business),
        key_trends=trends[:4],
        risks=risks[:3],
    )


def _margin_pct(info: dict) -> float | None:
    m = get_info_value(info, "profitMargins")
    if m is None:
        return None
    return m * 100 if abs(m) <= 1 else m


def _fetch_peer_metrics(ticker: str) -> PeerMetrics | None:
    try:
        stock = yf.Ticker(ticker)
        info = stock.info or {}
        if not info.get("symbol") and not info.get("shortName"):
            hist = stock.history(period="5d")
            if hist.empty:
                return None
        rev_growth = get_info_value(info, "revenueGrowth")
        if rev_growth is not None:
            rev_growth = rev_growth * 100 if abs(rev_growth) <= 1 else rev_growth
        return PeerMetrics(
            ticker=ticker.upper(),
            name=get_info_value(info, "shortName", "longName", default=ticker) or ticker,
            market_cap=get_info_value(info, "marketCap"),
            pe_ratio=get_info_value(info, "trailingPE"),
            ev_ebitda=get_info_value(info, "enterpriseToEbitda"),
            profit_margin=_margin_pct(info),
            revenue_growth=rev_growth,
            roe=get_info_value(info, "returnOnEquity"),
        )
    except Exception:
        return None


def _median(values: list[float | None]) -> float | None:
    clean = [v for v in values if v is not None and v == v]
    if not clean:
        return None
    clean.sort()
    n = len(clean)
    mid = n // 2
    return clean[mid] if n % 2 else (clean[mid - 1] + clean[mid]) / 2


def build_competitor_analysis(
    data: CompanyData,
    summary: FinancialSummary,
    peer_tickers: list[str] | None = None,
    max_peers: int = 4,
) -> CompetitorAnalysis:
    symbol = data.ticker
    if peer_tickers:
        peers_list = [p.upper() for p in peer_tickers if p.upper() != symbol][:max_peers]
    else:
        peers_list = discover_sector_peers(symbol, summary.sector, summary.industry, max_peers=max_peers)

    peers: list[PeerMetrics] = []
    for p in peers_list:
        metrics = _fetch_peer_metrics(p)
        if metrics:
            peers.append(metrics)

    target_peer = PeerMetrics(
        ticker=symbol,
        name=summary.company_name[:20],
        market_cap=summary.market_cap,
        pe_ratio=summary.pe_ratio,
        ev_ebitda=summary.ev_ebitda,
        profit_margin=summary.profit_margin,
        revenue_growth=summary.revenue_growth,
        roe=summary.roe,
    )

    all_peers = peers + [target_peer]
    med_pe = _median([p.pe_ratio for p in all_peers])
    med_ev = _median([p.ev_ebitda for p in all_peers])
    med_margin = _median([p.profit_margin for p in all_peers])

    position_parts: list[str] = []
    if summary.pe_ratio and med_pe:
        if summary.pe_ratio < med_pe * 0.85:
            position_parts.append("trades at a discount on P/E vs peers")
        elif summary.pe_ratio > med_pe * 1.15:
            position_parts.append("commands premium P/E vs peers")
        else:
            position_parts.append("P/E in line with peer median")

    if summary.profit_margin and med_margin:
        if summary.profit_margin > med_margin * 1.1:
            position_parts.append("stronger profitability than peers")
        elif summary.profit_margin < med_margin * 0.9:
            position_parts.append("below-peer profitability")

    relative = "; ".join(position_parts) if position_parts else "Peer comparison limited by data availability."

    return CompetitorAnalysis(
        target_ticker=symbol,
        peers=peers,
        peer_median_pe=med_pe,
        peer_median_ev_ebitda=med_ev,
        peer_median_margin=med_margin,
        relative_position=relative.capitalize(),
    )


def _estimate_cost_of_equity(info: dict, override: float | None = None) -> float:
    """CAPM cost of common equity (Ke), per Simon DCF sheet cell B3."""
    if override is not None:
        return override
    beta = get_info_value(info, "beta") or 1.0
    rf = 0.04
    erp = 0.055
    return float(min(max(rf + beta * erp, 0.075), 0.18))


def _estimate_wacc(info: dict, override: float | None = None) -> float:
    return _estimate_cost_of_equity(info, override)


def _compute_fcfe_for_year(cashflow: pd.DataFrame, balance_sheet: pd.DataFrame | None, col_index: int) -> float | None:
    """
    Free cash flow to common equity (Simon model DCF sheet):
    FCFE = Cash from operations + Cash from investing + Increase in debt.
    """
    ocf = _safe_series_value(cashflow, "Operating Cash Flow", col_index)
    if ocf is None:
        return None
    inv_cf = _safe_series_value(cashflow, "Investing Cash Flow", col_index)
    if inv_cf is None:
        capex = _safe_series_value(cashflow, "Capital Expenditure", col_index)
        inv_cf = capex if capex is not None else 0.0
    debt_chg = 0.0
    if balance_sheet is not None and not balance_sheet.empty:
        debt_now = _safe_series_value(balance_sheet, "Total Debt", col_index)
        debt_prev = _safe_series_value(balance_sheet, "Total Debt", col_index + 1)
        if debt_now is not None and debt_prev is not None:
            debt_chg = debt_now - debt_prev
    return ocf + inv_cf + debt_chg


def _standard_fcf_for_year(cashflow: pd.DataFrame, col_index: int) -> float | None:
    fcf = _safe_series_value(cashflow, "Free Cash Flow", col_index)
    if fcf is not None:
        return fcf
    ocf = _safe_series_value(cashflow, "Operating Cash Flow", col_index)
    capex = _safe_series_value(cashflow, "Capital Expenditure", col_index)
    if ocf is not None and capex is not None:
        return ocf + capex if capex < 0 else ocf - capex
    return None


def _historical_fcfe_series(data: CompanyData, max_years: int = 4) -> list[float]:
    cf = data.cashflow
    if cf is None or cf.empty:
        return []
    n = min(max_years, cf.shape[1])
    series: list[float] = []
    for col_i in range(n - 1, -1, -1):
        simon = _compute_fcfe_for_year(cf, data.balance_sheet, col_i)
        standard = _standard_fcf_for_year(cf, col_i)
        if simon is not None and standard is not None and standard > 0:
            if simon < standard * 0.5:
                v = standard
            else:
                v = simon
        else:
            v = simon if simon is not None else standard
        if v is not None and v == v:
            series.append(v)
    return series


def _simon_base_fcfe(data: CompanyData, summary: FinancialSummary | None = None) -> float | None:
    """Base FCFE: average historical (Simon), blended with TTM FCF and revenue-based proxy."""
    historical = _historical_fcfe_series(data)
    fcf_ttm = get_free_cash_flow(data.cashflow)
    candidates: list[float] = []
    if historical:
        candidates.append(sum(historical) / len(historical))
    if fcf_ttm and fcf_ttm > 0:
        candidates.append(fcf_ttm)
    if summary and summary.revenue and fcf_ttm and summary.revenue > 0:
        fcf_yield = fcf_ttm / summary.revenue
        if fcf_yield > 0:
            candidates.append(summary.revenue * fcf_yield)
    if not candidates:
        cf = data.cashflow
        if cf is None or cf.empty:
            return None
        return _compute_fcfe_for_year(cf, data.balance_sheet, 0) or _standard_fcf_for_year(cf, 0)
    return max(candidates)


def _fcfe_growth_rate(historical: list[float]) -> float:
    if len(historical) < 2:
        return DEFAULT_FCF_GROWTH
    start, end = historical[0], historical[-1]
    if start <= 0 or end <= 0:
        if start < 0 < end:
            return 0.08
        return DEFAULT_FCF_GROWTH
    n = len(historical) - 1
    try:
        cagr = (end / start) ** (1 / n) - 1
    except (ValueError, ZeroDivisionError):
        return DEFAULT_FCF_GROWTH
    return float(min(max(cagr, -0.03), 0.12))


def _terminal_growth_rate(info: dict, override: float | None = None) -> float:
    if override is not None:
        return min(override, 0.04)
    roe = get_info_value(info, "returnOnEquity")
    payout = get_info_value(info, "payoutRatio")
    if roe is not None and payout is not None:
        r = roe if abs(roe) <= 1 else roe / 100
        p = payout if abs(payout) <= 1 else payout / 100
        return float(min(max(r * (1 - min(max(p, 0), 1)), 0.015), 0.04))
    return DEFAULT_TERMINAL_GROWTH


def run_dcf(
    data: CompanyData,
    summary: FinancialSummary,
    wacc: float | None = None,
    terminal_growth: float | None = None,
    projection_years: int | None = None,
    fcf_growth: float | None = None,
) -> DCFResult:
    """
    Simon valuation model (FCFE): discount projected FCFE at cost of equity,
    terminal value = FCFE_n * (1+g) / (Ke-g), mid-year adjustment on equity value.
    """
    info = data.info
    ke = _estimate_cost_of_equity(info, wacc)
    years = projection_years or DEFAULT_PROJECTION_YEARS
    tg = _terminal_growth_rate(info, terminal_growth)
    tg = min(tg, ke - 0.01)

    historical = _historical_fcfe_series(data)
    base_fcfe = _simon_base_fcfe(data, summary)
    growth = fcf_growth if fcf_growth is not None else _fcfe_growth_rate(historical)
    rev_g = (summary.revenue_growth or 0) / 100
    growth = max(min(growth, 0.15), rev_g * 0.65, DEFAULT_TERMINAL_GROWTH)
    if summary.revenue_growth and summary.revenue_growth > 5:
        growth = max(growth, min(summary.revenue_growth / 100 * 0.85, 0.12))

    projected: list[float] = []
    pv_by_year: list[float] = []
    equity_value = None
    equity_value_raw = None
    implied_price = None
    pv_explicit = None
    pv_terminal = None
    terminal_fcfe_val = None
    shares = get_shares_outstanding(info)

    if base_fcfe is not None and base_fcfe == base_fcfe and ke > tg:
        fcfe = base_fcfe
        for _ in range(years):
            fcfe *= 1 + growth
            projected.append(fcfe)

        for t, f in enumerate(projected):
            pv_by_year.append(f / (1 + ke) ** (t + 1))
        pv_explicit = sum(pv_by_year)
        terminal_fcfe_val = projected[-1] * (1 + tg)
        pv_terminal = (terminal_fcfe_val / (ke - tg)) / (1 + ke) ** years
        equity_value_raw = pv_explicit + pv_terminal
        equity_value = equity_value_raw * (1 + ke / 2)

        if shares and shares > 0:
            implied_price = equity_value / shares

    current = summary.current_price
    upside = None
    if implied_price and current and current > 0:
        upside = (implied_price / current - 1) * 100

    return DCFResult(
        fcf_base=base_fcfe,
        wacc=ke,
        terminal_growth=tg,
        projection_years=years,
        fcf_growth=growth,
        enterprise_value=None,
        equity_value=equity_value,
        implied_price=implied_price,
        current_price=current,
        upside_pct=upside,
        projected_fcfe=projected,
        pv_by_year=pv_by_year,
        pv_explicit=pv_explicit,
        pv_terminal=pv_terminal,
        terminal_fcfe=terminal_fcfe_val,
        equity_value_raw=equity_value_raw,
        shares_used=shares,
        assumptions_note=f"Simon FCFE, Ke={ke:.1%}, g={tg:.1%}, FCFE growth={growth:.1%}, {years}Y",
    )


def run_comps(
    data: CompanyData,
    summary: FinancialSummary,
    comp_analysis: CompetitorAnalysis,
) -> CompsResult:
    info = data.info
    shares = get_shares_outstanding(info)
    current = summary.current_price

    med_pe = comp_analysis.peer_median_pe
    med_ev_ebitda = comp_analysis.peer_median_ev_ebitda

    eps = get_info_value(info, "trailingEps")
    ebitda = get_ebitda(data.financials, info)
    revenue = get_revenue(data.financials, info)
    net_debt = get_net_debt(data.balance_sheet) or 0

    implied_pe = med_pe * eps if med_pe and eps and eps > 0 else None

    implied_ev_ebitda = None
    if med_ev_ebitda and ebitda and ebitda > 0 and shares and shares > 0:
        implied_ev = med_ev_ebitda * ebitda
        implied_equity = implied_ev - net_debt
        implied_ev_ebitda = implied_equity / shares

    peer_ps_values = []
    for p in comp_analysis.peers:
        try:
            pi = yf.Ticker(p.ticker).info or {}
            ps = get_info_value(pi, "priceToSalesTrailing12Months")
            if ps:
                peer_ps_values.append(ps)
        except Exception:
            pass
    med_ps = None
    if peer_ps_values:
        peer_ps_values.sort()
        n = len(peer_ps_values)
        med_ps = peer_ps_values[n // 2]

    implied_ps = None
    if med_ps and revenue and revenue > 0 and shares and shares > 0:
        implied_equity = med_ps * revenue
        implied_ps = implied_equity / shares

    prices = [p for p in (implied_pe, implied_ev_ebitda, implied_ps) if p and p > 0]
    blended = sum(prices) / len(prices) if prices else None

    upside = None
    if blended and current and current > 0:
        upside = (blended / current - 1) * 100

    return CompsResult(
        peer_median_pe=med_pe,
        peer_median_ev_ebitda=med_ev_ebitda,
        peer_median_ps=med_ps,
        implied_price_pe=implied_pe,
        implied_price_ev_ebitda=implied_ev_ebitda,
        implied_price_ps=implied_ps,
        blended_fair_value=blended,
        current_price=current,
        upside_pct=upside,
    )


def _assess_dcf_reliability(
    dcf_price: float | None,
    current: float,
    comps_price: float | None,
) -> tuple[bool, str]:
    if not dcf_price or dcf_price <= 0:
        return False, "Insufficient FCFE / cash flow data"
    if current > 0 and dcf_price < current * 0.35:
        return False, "DCF implied price < 35% of market (capex-heavy or distorted FCFE)"
    if comps_price and comps_price > 0 and dcf_price < comps_price * 0.45:
        return False, "DCF diverges >55% below comparable-implied value"
    if current > 0 and dcf_price > current * 2.5:
        return False, "DCF implied price > 2.5x market (check growth assumptions)"
    return True, ""


def _method_outlier_reason(price: float, other_prices: list[float], current: float) -> str:
    if not other_prices:
        return ""
    med = statistics.median(other_prices)
    if med <= 0:
        return ""
    ratio_med = price / med
    if ratio_med < OUTLIER_VS_MEDIAN_LOW:
        return f"Diverges below other methods ({ratio_med:.0%} of median)"
    if ratio_med > OUTLIER_VS_MEDIAN_HIGH:
        return f"Diverges above other methods ({ratio_med:.0%} of median)"
    if current > 0:
        ratio_cur = price / current
        if ratio_cur < OUTLIER_VS_CURRENT_LOW:
            return f"Implied price {ratio_cur:.0%} of current (too low)"
        if ratio_cur > OUTLIER_VS_CURRENT_HIGH:
            return f"Implied price {ratio_cur:.0%} of current (too high)"
    return ""


def _filter_valuation_methods(
    methods: dict[str, float],
    current: float,
) -> tuple[dict[str, float], dict[str, str]]:
    """Drop methods that disagree sharply with peers; iterate until stable."""
    active = dict(methods)
    excluded: dict[str, str] = {}
    for _ in range(4):
        if len(active) < 2:
            break
        removed_any = False
        for name in list(active.keys()):
            price = active[name]
            others = [p for n, p in active.items() if n != name]
            reason = _method_outlier_reason(price, others, current)
            if reason:
                excluded[name] = reason
                del active[name]
                removed_any = True
        if not removed_any:
            break
    return active, excluded


def build_valuation_summary(
    data: CompanyData,
    summary: FinancialSummary,
    comp_analysis: CompetitorAnalysis,
    wacc: float | None = None,
) -> ValuationSummary:
    dcf = run_dcf(data, summary, wacc=wacc)
    comps = run_comps(data, summary, comp_analysis)

    current = summary.current_price or 0
    dcf_t = dcf.implied_price if dcf.implied_price and dcf.implied_price > 0 else None
    comps_t = comps.blended_fair_value if comps.blended_fair_value and comps.blended_fair_value > 0 else None
    analyst_t = summary.analyst_target if summary.analyst_target and summary.analyst_target > 0 else None

    methods: dict[str, float] = {}
    if dcf_t:
        methods["DCF"] = dcf_t
    if comps_t:
        methods["Comps"] = comps_t
    if analyst_t:
        methods["Analyst"] = analyst_t

    active, excluded = _filter_valuation_methods(methods, current)

    if dcf_t:
        dcf_hard_ok, hard_reason = _assess_dcf_reliability(dcf_t, current, comps_t)
        if not dcf_hard_ok:
            excluded["DCF"] = hard_reason
            active.pop("DCF", None)

    dcf_reliable = "DCF" in active
    dcf_reason = excluded.get("DCF", "")
    comps_reliable = "Comps" in active
    analyst_reliable = "Analyst" in active

    base_weights = {"DCF": 0.28, "Comps": 0.44, "Analyst": 0.28}
    weights = {k: base_weights[k] for k in active}
    if not weights and methods:
        # fallback: median of all raw methods
        prices = list(methods.values())
        med = statistics.median(prices)
        target_price = med
        weights = {}
    else:
        total_w = sum(weights.values())
        weights = {k: v / total_w for k, v in weights.items()}
        target_price = sum(active[k] * weights[k] for k in active) if active else None

    if target_price and current:
        diff_pct = (target_price / current - 1) * 100
        if diff_pct > 15:
            rec = "BUY - Models suggest meaningful upside"
        elif diff_pct > 5:
            rec = "OVERWEIGHT - Modest upside to fair value"
        elif diff_pct > -5:
            rec = "HOLD - Fairly valued"
        elif diff_pct > -15:
            rec = "UNDERWEIGHT - Trading above model fair value"
        else:
            rec = "SELL - Significant downside vs models"
    else:
        rec = "HOLD - Insufficient valuation data"

    return ValuationSummary(
        dcf=dcf,
        comps=comps,
        recommendation=rec,
        target_price=target_price,
        dcf_target=dcf_t,
        comps_target=comps_t,
        analyst_target=analyst_t,
        dcf_reliable=dcf_reliable,
        dcf_exclusion_reason=dcf_reason,
        comps_reliable=comps_reliable,
        analyst_reliable=analyst_reliable,
        excluded_methods=excluded,
        blend_weights=weights,
    )


def _upside_str(price: float | None, current: float) -> str:
    if not price or not current or current <= 0:
        return "N/A"
    return _fmt_pct((price / current - 1) * 100, signed=True)


def build_valuation_crosscheck_table(
    valuation: ValuationSummary,
    summary: FinancialSummary,
) -> list[list[str]]:
    """Multi-method valuation comparison for cross-validation."""
    current = summary.current_price or 0
    comps = valuation.comps
    dcf = valuation.dcf
    rows: list[list[str]] = [
        ["Valuation Method", "Implied Price", "vs. Current", "Weight in Target", "Notes"],
    ]

    def _weight_cell(method: str) -> str:
        if method in valuation.blend_weights:
            return f"{valuation.blend_weights[method]:.0%}"
        return "0% (excluded)"

    def _note_cell(method: str) -> str:
        if method in valuation.excluded_methods:
            return valuation.excluded_methods[method]
        if method in valuation.blend_weights:
            return "Included in blend"
        return "N/A"

    rows.append([
        "DCF (Simon FCFE)",
        f"${valuation.dcf_target:.2f}" if valuation.dcf_target else "N/A",
        _upside_str(valuation.dcf_target, current),
        _weight_cell("DCF"),
        _note_cell("DCF"),
    ])

    if comps.implied_price_pe:
        rows.append([
            "Comps — P/E (peer median)",
            f"${comps.implied_price_pe:.2f}",
            _upside_str(comps.implied_price_pe, current),
            "—",
            f"Peer median P/E {_fmt_ratio(comps.peer_median_pe)}",
        ])
    if comps.implied_price_ev_ebitda:
        rows.append([
            "Comps — EV/EBITDA (peer median)",
            f"${comps.implied_price_ev_ebitda:.2f}",
            _upside_str(comps.implied_price_ev_ebitda, current),
            "—",
            f"Peer median {_fmt_ratio(comps.peer_median_ev_ebitda)}",
        ])
    if comps.implied_price_ps:
        rows.append([
            "Comps — P/S (peer median)",
            f"${comps.implied_price_ps:.2f}",
            _upside_str(comps.implied_price_ps, current),
            "—",
            f"Peer median P/S {_fmt_ratio(comps.peer_median_ps)}",
        ])
    rows.append([
        "Comps — Blended (avg. of available multiples)",
        f"${valuation.comps_target:.2f}" if valuation.comps_target else "N/A",
        _upside_str(valuation.comps_target, current),
        _weight_cell("Comps"),
        _note_cell("Comps") if not valuation.comps_reliable else "Anchor multiple-based value",
    ])
    rows.append([
        "Analyst consensus (Yahoo Finance)",
        f"${valuation.analyst_target:.2f}" if valuation.analyst_target else "N/A",
        _upside_str(valuation.analyst_target, current),
        _weight_cell("Analyst"),
        _note_cell("Analyst"),
    ])
    rows.append([
        "Blended target (weighted)",
        f"${valuation.target_price:.2f}" if valuation.target_price else "N/A",
        _upside_str(valuation.target_price, current),
        "100%",
        " + ".join(f"{k} {v:.0%}" for k, v in valuation.blend_weights.items()) or "Comps only",
    ])
    return rows


def recommendation_short(valuation: ValuationSummary) -> str:
    rec = valuation.recommendation.upper()
    for label in ("BUY", "OVERWEIGHT", "HOLD", "UNDERWEIGHT", "SELL"):
        if label in rec:
            return label
    return "HOLD"


def build_executive_summary(summary: FinancialSummary, industry: IndustryResearch, valuation: ValuationSummary) -> str:
    name = summary.company_name
    sector = summary.sector
    rating = recommendation_short(valuation)
    target = f"${valuation.target_price:.2f}" if valuation.target_price else "N/A"
    price = f"${summary.current_price:.2f}" if summary.current_price else "N/A"
    upside = ""
    if valuation.target_price and summary.current_price and summary.current_price > 0:
        pct = (valuation.target_price / summary.current_price - 1) * 100
        upside = f", implying {pct:+.1f}% upside from the current price of {price}"
    cross = (
        " Valuation cross-check uses DCF (Simon FCFE), comparable multiples, and analyst consensus; "
        "methods that diverge materially from peers are down-weighted."
    )
    thesis = (
        f"{name} operates in {sector} ({summary.industry}). "
        f"{industry.business_summary} "
        f"Our multi-method valuation supports a <b>{rating}</b> rating with a blended target of {target}{upside}.{cross}"
    )
    return thesis


def build_financial_health_summary(summary: FinancialSummary) -> str:
    notes: list[str] = []
    if summary.revenue_growth is not None:
        if summary.revenue_growth > 10:
            notes.append("strong revenue growth")
        elif summary.revenue_growth > 0:
            notes.append("modest revenue growth")
        else:
            notes.append("declining revenue")
    if summary.operating_margin is not None:
        if summary.operating_margin > 20:
            notes.append("high operating margins")
        elif summary.operating_margin > 10:
            notes.append("healthy margins")
        else:
            notes.append("compressed margins")
    if summary.debt_to_equity and summary.debt_to_equity > 150:
        notes.append("elevated leverage")
    elif summary.debt_to_equity and summary.debt_to_equity < 50:
        notes.append("conservative balance sheet")
    if summary.roe is not None:
        if summary.roe > 20:
            notes.append("strong ROE")
        elif summary.roe < 5:
            notes.append("weak ROE")
    if summary.free_cash_flow and summary.free_cash_flow > 0:
        notes.append("positive FCF generation")
    if not notes:
        return "Mixed financial profile; monitor sector and macro drivers."
    return "; ".join(notes).capitalize() + "."


def build_multi_year_financial_rows(data: CompanyData) -> list[list[str]]:
    fin = data.financials
    if fin is None or fin.empty:
        return []
    years: list[str] = []
    rows_map: dict[str, list[str]] = {
        "Revenue": [],
        "Net Income": [],
        "Gross Margin": [],
        "Net Margin": [],
    }
    label_map = {
        "Revenue": "Total Revenue",
        "Net Income": "Net Income",
    }
    max_cols = min(3, fin.shape[1])
    for col_i in range(max_cols):
        try:
            dt = fin.columns[col_i]
            years.append(str(dt.year) if hasattr(dt, "year") else str(dt)[:4])
        except Exception:
            years.append(f"Y{col_i}")

    for display, row_key in label_map.items():
        vals = []
        for col_i in range(max_cols):
            v = _safe_series_value(fin, row_key, col_i)
            vals.append(format_large_number(v) if v is not None else "N/A")
        rows_map[display] = vals

    for col_i in range(max_cols):
        rev = _safe_series_value(fin, "Total Revenue", col_i)
        gp = _safe_series_value(fin, "Gross Profit", col_i)
        ni = _safe_series_value(fin, "Net Income", col_i)
        if rev and rev != 0 and gp is not None:
            rows_map["Gross Margin"].append(f"{gp / rev * 100:.1f}%")
        else:
            rows_map["Gross Margin"].append("N/A")
        if rev and rev != 0 and ni is not None:
            rows_map["Net Margin"].append(f"{ni / rev * 100:.1f}%")
        else:
            rows_map["Net Margin"].append("N/A")

    header = ["Metric"] + years
    table_rows = [header]
    for metric in ("Revenue", "Net Income", "Gross Margin", "Net Margin"):
        table_rows.append([metric] + rows_map[metric])
    return table_rows


def _year_label(fin: pd.DataFrame, col_index: int, suffix: str) -> str:
    try:
        dt = fin.columns[col_index]
        yr = dt.year if hasattr(dt, "year") else str(dt)[:4]
        return f"FY {yr}{suffix}"
    except Exception:
        return f"FY Y{col_index}{suffix}"


def _fmt_fin_num(value: float | None, in_millions: bool = False) -> str:
    if value is None:
        return "N/A"
    if in_millions:
        return f"{value / 1e6:,.0f}"
    return f"{value:,.2f}"


def build_financial_valuation_summary_table(
    data: CompanyData,
    summary: FinancialSummary,
) -> list[list[str]]:
    """
    Broker-style financial & valuation summary: 2 historical years (A) + 3 forecast years (E).
    """
    fin = data.financials
    bs = data.balance_sheet
    info = data.info
    shares = get_shares_outstanding(info) or 1.0
    price = summary.current_price

    if fin is None or fin.empty:
        return []

    hist_cols = min(2, fin.shape[1])
    headers = [_year_label(fin, hist_cols - 1 - i, "A") for i in range(hist_cols)]
    try:
        newest_yr = fin.columns[0].year if hasattr(fin.columns[0], "year") else datetime.now().year
    except Exception:
        newest_yr = datetime.now().year
    for i in range(3):
        headers.append(f"FY {newest_yr + i + 1}E")

    def _hist(metric: str, col: int) -> float | None:
        return _safe_series_value(fin, metric, col)

    revenues_h: list[float | None] = [_hist("Total Revenue", i) for i in range(hist_cols)]
    net_income_h: list[float | None] = [_hist("Net Income", i) for i in range(hist_cols)]

    base_rev = revenues_h[0] if revenues_h and revenues_h[0] else summary.revenue
    if summary.revenue_growth is not None:
        g0 = summary.revenue_growth / 100
    elif len(revenues_h) >= 2 and revenues_h[0] and revenues_h[1]:
        g0 = (revenues_h[0] - revenues_h[1]) / abs(revenues_h[1])
    else:
        g0 = 0.05

    revenues_f: list[float] = []
    net_income_f: list[float] = []
    rev = base_rev or 0
    ni_margin = None
    if net_income_h[0] and revenues_h[0] and revenues_h[0] != 0:
        ni_margin = net_income_h[0] / revenues_h[0]
    elif summary.profit_margin:
        ni_margin = summary.profit_margin / 100

    for i in range(3):
        fade = 0.7 ** i
        g = g0 * fade + DEFAULT_TERMINAL_GROWTH * (1 - fade)
        rev *= 1 + g
        revenues_f.append(rev)
        net_income_f.append(rev * ni_margin if ni_margin else 0)

    all_rev = list(reversed(revenues_h)) + revenues_f
    all_ni = list(reversed(net_income_h)) + net_income_f

    rev_growth: list[str] = []
    for i, r in enumerate(all_rev):
        if i == 0:
            rev_growth.append("N/A")
        elif all_rev[i - 1] and all_rev[i - 1] != 0 and r is not None:
            rev_growth.append(f"{(r - all_rev[i - 1]) / abs(all_rev[i - 1]) * 100:.1f}%")
        else:
            rev_growth.append("N/A")

    ni_growth: list[str] = []
    for i, ni in enumerate(all_ni):
        if i == 0:
            ni_growth.append("N/A")
        elif all_ni[i - 1] and all_ni[i - 1] != 0 and ni is not None:
            ni_growth.append(f"{(ni - all_ni[i - 1]) / abs(all_ni[i - 1]) * 100:.1f}%")
        else:
            ni_growth.append("N/A")

    gross_margins: list[str] = []
    net_margins: list[str] = []
    for i in range(5):
        if i < hist_cols:
            col = hist_cols - 1 - i
            rev_v = _hist("Total Revenue", col)
            gp = _hist("Gross Profit", col)
            ni = _hist("Net Income", col)
        else:
            rev_v = all_rev[i]
            ni = all_ni[i]
            gp = rev_v * (summary.gross_margin / 100) if rev_v and summary.gross_margin else None
        if rev_v and rev_v != 0 and gp is not None:
            gross_margins.append(f"{gp / rev_v * 100:.1f}%")
        else:
            gross_margins.append("N/A")
        if rev_v and rev_v != 0 and ni is not None:
            net_margins.append(f"{ni / rev_v * 100:.1f}%")
        else:
            net_margins.append("N/A")

    roe_vals: list[str] = []
    eps_vals: list[str] = []
    bps_vals: list[str] = []
    pe_vals: list[str] = []
    pb_vals: list[str] = []
    div_vals: list[str] = []

    for i in range(5):
        if i < hist_cols:
            col = hist_cols - 1 - i
            ni = _hist("Net Income", col)
            eq = _safe_series_value(bs, "Stockholders Equity", col) if bs is not None else None
        else:
            ni = all_ni[i]
            eq = _safe_series_value(bs, "Stockholders Equity", 0) if bs is not None else None
            if eq and i > hist_cols:
                eq *= 1.05 ** (i - hist_cols + 1)
        eps = ni / shares if ni is not None else None
        bps = eq / shares if eq is not None else None
        eps_vals.append(_fmt_fin_num(eps))
        bps_vals.append(_fmt_fin_num(bps))
        if eq and ni is not None and eq != 0:
            roe_vals.append(f"{ni / eq * 100:.1f}%")
        elif summary.roe and i >= hist_cols:
            roe_vals.append(_fmt_pct(summary.roe))
        else:
            roe_vals.append("N/A")
        if price and eps and eps > 0:
            pe_vals.append(f"{price / eps:.2f}")
        else:
            pe_vals.append("N/A")
        if price and bps and bps > 0:
            pb_vals.append(f"{price / bps:.2f}")
        else:
            pb_vals.append("N/A")
        div_vals.append(_fmt_pct(summary.dividend_yield) if summary.dividend_yield else "N/A")

    rows = [
        ["Metric"] + headers,
        ["Revenue ($M)"] + [_fmt_fin_num(r, True) if r else "N/A" for r in all_rev],
        ["Revenue Growth (%)"] + rev_growth,
        ["Net Income ($M)"] + [_fmt_fin_num(n, True) if n else "N/A" for n in all_ni],
        ["Net Income Growth (%)"] + ni_growth,
        ["Gross Margin (%)"] + gross_margins,
        ["Net Margin (%)"] + net_margins,
        ["ROE (%)"] + roe_vals,
        ["EPS ($)"] + eps_vals,
        ["Book Value / Share ($)"] + bps_vals,
        ["P/E (x)"] + pe_vals,
        ["P/B (x)"] + pb_vals,
        ["Dividend Yield (%)"] + div_vals,
    ]
    return rows


def fetch_top_us_market_cap(limit: int = TOP_US_SUMMARY_COUNT) -> list[str]:
    """Return the largest US-listed companies by market cap from a mega-cap universe."""
    candidates: list[tuple[str, float]] = []
    seen_names: set[str] = set()
    for i, ticker in enumerate(MEGA_CAP_UNIVERSE):
        if i > 0:
            time.sleep(UNIVERSE_FETCH_DELAY_SEC)
        try:
            info = yf.Ticker(ticker).info or {}
            mcap = info.get("marketCap")
            if not mcap:
                continue
            country = (info.get("country") or "").lower()
            if country and country not in ("united states", "usa"):
                continue
            name = (info.get("shortName") or info.get("longName") or ticker).lower()
            # Prefer GOOGL over GOOG when both appear.
            dedupe_key = name.split()[0] if name else ticker
            if dedupe_key in seen_names and ticker in ("GOOG", "BRK-A"):
                continue
            seen_names.add(dedupe_key)
            candidates.append((ticker.upper(), float(mcap)))
        except Exception:
            continue
    candidates.sort(key=lambda x: x[1], reverse=True)
    return [t for t, _ in candidates[:limit]]


def clean_report_directory(output_dir: str, keep_summary: bool = False) -> None:
    if not os.path.isdir(output_dir):
        return
    for pattern in ("*.pdf", "*.csv"):
        for path in glob.glob(os.path.join(output_dir, pattern)):
            if keep_summary and os.path.basename(path) in (
                SUMMARY_REPORT_NAME, "Top20_Summary.csv", "Top50_Summary.pdf", "Top50_Summary.csv",
            ):
                continue
            try:
                os.remove(path)
            except OSError:
                pass
    for legacy in ("Top50_Summary.pdf", "Top50_Summary.csv"):
        path = os.path.join(output_dir, legacy)
        if os.path.isfile(path):
            try:
                os.remove(path)
            except OSError:
                pass


def reset_data_state(output_dir: str) -> None:
    path = data_state_path(output_dir)
    if os.path.isfile(path):
        try:
            os.remove(path)
        except OSError:
            pass


def is_monday(d: datetime | None = None) -> bool:
    d = d or datetime.now()
    return d.weekday() == 0


def _analyze_ticker(
    ticker: str,
    peer_tickers: list[str] | None = None,
    wacc: float | None = None,
) -> tuple[CompanyData, FinancialSummary, IndustryResearch, CompetitorAnalysis, ValuationSummary] | None:
    symbol = ticker.upper().strip()
    try:
        data = fetch_company(symbol)
    except ValueError:
        print(f"  Skipping {symbol}: data unavailable", file=sys.stderr)
        return None
    except Exception as e:
        print(f"  Skipping {symbol}: {e}", file=sys.stderr)
        return None
    summary = build_financial_summary(data)
    industry = build_industry_research(data, summary)
    if peer_tickers is None:
        top_peers = fetch_top_us_market_cap(TOP_US_SUMMARY_COUNT)
        peer_tickers = [p for p in top_peers if p != symbol][:4]
    competitors = build_competitor_analysis(data, summary, peer_tickers=peer_tickers)
    valuation = build_valuation_summary(data, summary, competitors, wacc=wacc)
    return data, summary, industry, competitors, valuation


def _report_row_from_analysis(
    symbol: str,
    summary: FinancialSummary,
    valuation: ValuationSummary,
    output_path: str = "",
) -> ReportRow:
    upside = None
    if valuation.target_price and summary.current_price and summary.current_price > 0:
        upside = (valuation.target_price / summary.current_price - 1) * 100
    return ReportRow(
        ticker=symbol,
        company_name=summary.company_name,
        sector=summary.sector,
        current_price=summary.current_price,
        market_cap=summary.market_cap,
        pe_ratio=summary.pe_ratio,
        revenue_growth=summary.revenue_growth,
        operating_margin=summary.operating_margin,
        roe=summary.roe,
        dcf_fair_value=valuation.dcf.implied_price,
        comps_fair_value=valuation.comps.blended_fair_value,
        target_price=valuation.target_price,
        upside_pct=upside,
        rating=recommendation_short(valuation),
        financial_health=build_financial_health_summary(summary),
        output_path=output_path,
    )


def process_ticker_report(
    ticker: str,
    output_dir: str,
    peer_tickers: list[str] | None = None,
    wacc: float | None = None,
    replace: bool = True,
    data: CompanyData | None = None,
) -> tuple[ReportRow | None, CompanyData | None]:
    if data is None:
        result = _analyze_ticker(ticker, peer_tickers, wacc)
        if not result:
            return None, None
        data, summary, industry, competitors, valuation = result
    else:
        summary = build_financial_summary(data)
        industry = build_industry_research(data, summary)
        if peer_tickers is None:
            top_peers = fetch_top_us_market_cap(TOP_US_SUMMARY_COUNT)
            peer_tickers = [p for p in top_peers if p != data.ticker][:4]
        competitors = build_competitor_analysis(data, summary, peer_tickers=peer_tickers)
        valuation = build_valuation_summary(data, summary, competitors, wacc=wacc)

    symbol = ticker.upper().strip()

    os.makedirs(output_dir, exist_ok=True)
    filename = f"{symbol}_research.pdf" if replace else (
        f"{symbol}_research_{datetime.now().strftime('%Y%m%d')}.pdf"
    )
    output_path = os.path.join(output_dir, filename)
    if os.path.exists(output_path):
        try:
            os.remove(output_path)
        except OSError:
            base, ext = os.path.splitext(output_path)
            output_path = f"{base}_{datetime.now().strftime('%H%M%S')}{ext}"

    generate_pdf_report(summary, industry, competitors, valuation, data.history, output_path, data)
    row = _report_row_from_analysis(symbol, summary, valuation, output_path)
    return row, data


def process_ticker_summary_only(
    ticker: str,
    peer_tickers: list[str] | None = None,
    wacc: float | None = None,
    data: CompanyData | None = None,
) -> tuple[ReportRow | None, CompanyData | None]:
    if data is None:
        result = _analyze_ticker(ticker, peer_tickers, wacc)
        if not result:
            return None, None
        data, summary, _, _, valuation = result
    else:
        summary = build_financial_summary(data)
        if peer_tickers is None:
            top_peers = fetch_top_us_market_cap(TOP_US_SUMMARY_COUNT)
            peer_tickers = [p for p in top_peers if p != data.ticker][:4]
        competitors = build_competitor_analysis(data, summary, peer_tickers=peer_tickers)
        valuation = build_valuation_summary(data, summary, competitors, wacc=wacc)
    row = _report_row_from_analysis(ticker.upper().strip(), summary, valuation)
    return row, data


def build_pdf_ticker_list(ranked_tickers: list[str], pdf_count: int) -> list[str]:
    """Top N by market cap plus mandatory extra PDF tickers (e.g. TSLA, BX)."""
    pdf_tickers: list[str] = []
    for t in ranked_tickers[:pdf_count]:
        if t not in pdf_tickers:
            pdf_tickers.append(t)
    for t in EXTRA_PDF_TICKERS:
        if t not in pdf_tickers:
            pdf_tickers.append(t)
    return pdf_tickers


def _batch_process_one(
    ticker: str,
    output_dir: str,
    state: dict[str, Any],
    wacc: float | None,
    replace: bool,
    make_pdf: bool,
    refresh_stale: bool,
    force: bool,
    max_age_days: int = STALE_MAX_AGE_DAYS,
    allow_cache: bool = True,
) -> ReportRow | None:
    """Process one ticker; honor --refresh-stale unless --force."""
    symbol = ticker.upper()
    data: CompanyData | None = None
    if refresh_stale and not force:
        try:
            data = fetch_company(symbol)
        except Exception as e:
            print(f"  {symbol}: fetch failed ({e}); will retry via full pipeline", file=sys.stderr)
        if data is not None:
            refresh, reason = confirm_ticker_refresh(
                symbol, state, data, force=False, max_age_days=max_age_days,
            )
            if not refresh:
                print(f"  Skip {symbol} ({reason})")
                if allow_cache:
                    return _cached_report_row(state, symbol)
                return None
            print(f"  Refresh {symbol}: {reason}")

    for attempt in range(2):
        if make_pdf:
            row, data = process_ticker_report(
                symbol, output_dir, wacc=wacc, replace=replace, data=data if attempt == 0 else None,
            )
        else:
            row, data = process_ticker_summary_only(
                symbol, wacc=wacc, data=data if attempt == 0 else None,
            )
        if row and data:
            record_ticker_state(state, symbol, data, row)
            return row
        if attempt == 0:
            print(f"  Retry {symbol} after failure...", file=sys.stderr)
            data = None
            time.sleep(BATCH_TICKER_DELAY_SEC * 2)

    if allow_cache and refresh_stale and not force:
        return _cached_report_row(state, symbol)
    return None


def run_batch(
    output_dir: str = DEFAULT_OUTPUT_DIR,
    wacc: float | None = None,
    replace_old: bool = True,
    pdf_count: int = TOP_US_PDF_REPORT_COUNT,
    summary_count: int = TOP_US_SUMMARY_COUNT,
    refresh_stale: bool = False,
    force: bool = False,
    max_age_days: int = STALE_MAX_AGE_DAYS,
) -> list[ReportRow]:
    os.makedirs(output_dir, exist_ok=True)
    state = load_data_state(output_dir)

    if replace_old and not refresh_stale:
        clean_report_directory(output_dir)
        reset_data_state(output_dir)
        state = {"tickers": {}, "summary_last_run": None}

    summary_tickers = fetch_top_us_market_cap(summary_count)
    if len(summary_tickers) < summary_count:
        print(
            f"Warning: ranked only {len(summary_tickers)} tickers (requested {summary_count}).",
            file=sys.stderr,
        )
    pdf_tickers = build_pdf_ticker_list(summary_tickers, pdf_count)
    pdf_only = [t for t in pdf_tickers if t not in summary_tickers]
    all_work = list(dict.fromkeys(summary_tickers + pdf_only))

    mode = "refresh-stale" if refresh_stale and not force else "full"
    print(f"Top {summary_count} for summary: {', '.join(summary_tickers)}")
    print(f"PDF reports ({len(pdf_tickers)}): {', '.join(pdf_tickers)}")
    print(f"Batch mode: {mode}")
    if pdf_only:
        print(f"PDF-only (outside top {summary_count} summary): {', '.join(pdf_only)}")

    allow_cache = refresh_stale and not force
    rows_by_ticker: dict[str, ReportRow] = {}
    updated = 0
    for i, ticker in enumerate(summary_tickers, 1):
        print(f"[{i}/{len(summary_tickers)}] {ticker}...", flush=True)
        row = _batch_process_one(
            ticker, output_dir, state, wacc, replace_old,
            make_pdf=ticker in pdf_tickers,
            refresh_stale=refresh_stale, force=force, max_age_days=max_age_days,
            allow_cache=allow_cache,
        )
        if row:
            rows_by_ticker[ticker] = row
            updated += 1
        if i < len(summary_tickers):
            time.sleep(BATCH_TICKER_DELAY_SEC)

    for ticker in pdf_only:
        print(f"PDF (extra) {ticker}...", flush=True)
        _batch_process_one(
            ticker, output_dir, state, wacc, replace_old,
            make_pdf=True, refresh_stale=refresh_stale, force=force,
            max_age_days=max_age_days, allow_cache=allow_cache,
        )
        time.sleep(BATCH_TICKER_DELAY_SEC)

    missing = [t for t in summary_tickers if t not in rows_by_ticker]
    if missing and not allow_cache:
        print(f"Retrying {len(missing)} missing summary tickers...", flush=True)
        for ticker in missing:
            row = _batch_process_one(
                ticker, output_dir, state, wacc, replace_old,
                make_pdf=ticker in pdf_tickers,
                refresh_stale=False, force=True, max_age_days=max_age_days,
                allow_cache=False,
            )
            if row:
                rows_by_ticker[ticker] = row
                updated += 1
            time.sleep(BATCH_TICKER_DELAY_SEC)
        missing = [t for t in summary_tickers if t not in rows_by_ticker]

    rows = [rows_by_ticker[t] for t in summary_tickers if t in rows_by_ticker]
    required_rows = len(summary_tickers)
    if len(rows) >= required_rows:
        rows.sort(key=lambda r: r.market_cap or 0, reverse=True)
        summary_path = os.path.join(output_dir, SUMMARY_REPORT_NAME)
        generate_summary_pdf(rows, summary_path)
        state["summary_last_run"] = datetime.now(timezone.utc).isoformat()
        print(
            f"\nSummary report saved: {summary_path} "
            f"({len(rows)} companies, {updated} refreshed)",
        )
    else:
        print(
            f"\nSummary NOT updated: only {len(rows)}/{required_rows} companies "
            f"(missing: {', '.join(missing)}). Fix errors and re-run.",
            file=sys.stderr,
        )
    save_data_state(output_dir, state)
    return rows


def run_batch_top20(
    output_dir: str = DEFAULT_OUTPUT_DIR,
    wacc: float | None = None,
    replace_old: bool = True,
) -> list[ReportRow]:
    """Top 5 PDFs + TSLA/BX extras + Top 20 summary table."""
    return run_batch(output_dir, wacc, replace_old, TOP_US_PDF_REPORT_COUNT, TOP_US_SUMMARY_COUNT)


# ---------------------------------------------------------------------------
# PDF report
# ---------------------------------------------------------------------------


def _price_chart(history, ticker: str) -> io.BytesIO | None:
    if history is None or history.empty:
        return None
    fig, ax = plt.subplots(figsize=(3.2, 1.0), dpi=120)
    ax.plot(history.index, history["Close"], color="#2b6cb0", linewidth=1.2)
    ax.fill_between(history.index, history["Close"], alpha=0.15, color="#2b6cb0")
    ax.set_title(f"{ticker} — 1Y Price", fontsize=7, pad=3)
    ax.tick_params(labelsize=5)
    ax.grid(True, alpha=0.3, linewidth=0.5)
    fig.patch.set_facecolor("white")
    plt.tight_layout(pad=0.3)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf


def _cell(text: str, size: int = 7) -> Paragraph:
    return Paragraph(
        str(text),
        ParagraphStyle(name="Cell", fontName="Helvetica", fontSize=size, leading=size + 2, textColor=colors.black),
    )


def _note_cell(text: str) -> Paragraph:
    """Notes column: smaller font, wraps long exclusion / method notes."""
    return Paragraph(
        str(text),
        ParagraphStyle(
            name="NoteCell", fontName="Helvetica", fontSize=5.5, leading=7,
            textColor=colors.black, wordWrap="CJK",
        ),
    )


def _header_cell(text: str) -> Paragraph:
    """Section labels (navy text on white background)."""
    return Paragraph(
        str(text),
        ParagraphStyle(name="HeaderCell", fontName="Helvetica-Bold", fontSize=7, leading=9, textColor=NAVY),
    )


def _th_cell(text: str) -> Paragraph:
    """Table column headers (white text on navy background)."""
    return Paragraph(
        str(text),
        ParagraphStyle(name="ThCell", fontName="Helvetica-Bold", fontSize=6.5, leading=8, textColor=colors.white),
    )


def _metric_cell(text: str) -> Paragraph:
    """Row labels in the first column of data tables (dark text on light rows)."""
    return Paragraph(
        str(text),
        ParagraphStyle(name="MetricCell", fontName="Helvetica-Bold", fontSize=6, leading=8, textColor=NAVY),
    )


def build_dcf_detail_table_compact(dcf: DCFResult) -> list[list[str]]:
    """Shorter DCF table for one-page PDF layout."""
    rows = [
        ["Item", "Value"],
        ["Ke (cost of equity)", f"{dcf.wacc:.2%}"],
        ["Terminal g", f"{dcf.terminal_growth:.2%}"],
        ["FCFE growth", f"{dcf.fcf_growth:.2%}"],
        ["Base FCFE", format_large_number(dcf.fcf_base)],
        ["PV explicit (5Y)", format_large_number(dcf.pv_explicit)],
        ["PV terminal", format_large_number(dcf.pv_terminal)],
        ["Equity value (adj.)", format_large_number(dcf.equity_value)],
        ["Implied price", f"${dcf.implied_price:.2f}" if dcf.implied_price else "N/A"],
    ]
    if dcf.upside_pct is not None:
        rows.append(["vs. current", _fmt_pct(dcf.upside_pct, signed=True)])
    return rows


def build_dcf_detail_table(dcf: DCFResult) -> list[list[str]]:
    """Simon FCFE DCF walkthrough for PDF display."""
    rows: list[list[str]] = [
        ["Line Item", "Value / Formula"],
        ["Model", "Simon FCFE DCF (equity value to common shareholders)"],
        ["Cost of equity (Ke)", f"{dcf.wacc:.2%}"],
        ["Terminal growth (g)", f"{dcf.terminal_growth:.2%}"],
        ["FCFE growth (forecast)", f"{dcf.fcf_growth:.2%}"],
        ["Projection horizon", f"{dcf.projection_years} years"],
        ["Base FCFE (avg. historical)", format_large_number(dcf.fcf_base)],
    ]
    if not dcf.projected_fcfe:
        rows.append(["Status", "Insufficient data to project FCFE"])
        return rows

    rows.append(["", ""])
    rows.append(["Year", "Projected FCFE", "Discount factor 1/(1+Ke)^t", "Present value"])
    ke = dcf.wacc
    for t, (fcfe, pv) in enumerate(zip(dcf.projected_fcfe, dcf.pv_by_year), start=1):
        df = 1 / (1 + ke) ** t
        rows.append([f"Year {t}", format_large_number(fcfe), f"{df:.4f}", format_large_number(pv)])

    rows.extend([
        ["", "", "", ""],
        ["PV of explicit period (Y1-Y5)", "", "", format_large_number(dcf.pv_explicit)],
        [
            "Terminal-year FCFE",
            format_large_number(dcf.terminal_fcfe),
            f"= Y{dcf.projection_years} x (1+g)",
            "",
        ],
        [
            "Terminal value at Y5",
            format_large_number(
                (dcf.terminal_fcfe / (ke - dcf.terminal_growth)) if dcf.terminal_fcfe and ke > dcf.terminal_growth else None
            ),
            f"= FCFE / (Ke-g)",
            "",
        ],
        ["PV of terminal value", "", f"/ (1+Ke)^{dcf.projection_years}", format_large_number(dcf.pv_terminal)],
        ["Equity value (pre mid-year adj.)", format_large_number(dcf.equity_value_raw), "= PV explicit + PV terminal", ""],
        ["Mid-year convention adjustment", f"x (1 + Ke/2) = x {1 + ke / 2:.4f}", "", ""],
        ["Equity value (adjusted)", format_large_number(dcf.equity_value), "", ""],
        ["Shares outstanding", f"{dcf.shares_used / 1e6:.2f}M" if dcf.shares_used else "N/A", "", ""],
        ["Implied price per share", f"${dcf.implied_price:.2f}" if dcf.implied_price else "N/A", "", ""],
    ])
    if dcf.upside_pct is not None:
        rows.append(["vs. current price", _fmt_pct(dcf.upside_pct, signed=True), "", ""])
    return rows


def _section_title(text: str) -> Paragraph:
    return Paragraph(
        text.upper(),
        ParagraphStyle(
            name="Section", fontName="Helvetica-Bold", fontSize=7.5, leading=9,
            textColor=ACCENT, spaceBefore=1, spaceAfter=1,
        ),
    )


def _body(text: str, size: float = 6.5) -> Paragraph:
    return Paragraph(
        text,
        ParagraphStyle(name="Body", fontName="Helvetica", fontSize=size, leading=size + 2, textColor=colors.black),
    )


def _fmt_pct(v: float | None, signed: bool = False) -> str:
    if v is None:
        return "N/A"
    if signed and v > 0:
        return f"+{v:.1f}%"
    return f"{v:.1f}%"


def _fmt_ratio(v: float | None, suffix: str = "x") -> str:
    if v is None:
        return "N/A"
    return f"{v:.1f}{suffix}"


def _table_style_header() -> TableStyle:
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.25, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_BG]),
    ])


def _make_styled_table(
    rows: list[list[str]],
    col_widths: list[float],
    metric_label_col: bool = True,
    allow_split: bool = False,
    wrap_col: int | None = None,
) -> Table:
    """Build a table with correct header / row-label / data cell styles."""
    table_data: list[list] = []
    for ri, row in enumerate(rows):
        if ri == 0:
            table_data.append([_th_cell(str(c)) for c in row])
        elif metric_label_col and len(row) > 1:
            cells: list = [_metric_cell(row[0])]
            for col_i, val in enumerate(row[1:], start=1):
                cells.append(_note_cell(str(val)) if wrap_col == col_i else _cell(str(val)))
            table_data.append(cells)
        else:
            cells = []
            for col_i, val in enumerate(row):
                cells.append(_note_cell(str(val)) if wrap_col == col_i else _cell(str(val)))
            table_data.append(cells)
    tbl = Table(table_data, colWidths=col_widths)
    tbl.setStyle(_table_style_header())
    if not allow_split:
        tbl.splitByRow = 0
        tbl.splitInRow = 0
    return tbl


def build_financial_valuation_summary_table_pdf(
    data: CompanyData,
    summary: FinancialSummary,
) -> list[list[str]]:
    """Compact financial table for two-page PDF (fewer rows)."""
    full = build_financial_valuation_summary_table(data, summary)
    if not full:
        return full
    keep = {
        "Revenue ($M)", "Revenue Growth (%)", "Net Income ($M)",
        "Gross Margin (%)", "Net Margin (%)", "EPS ($)", "P/E (x)",
    }
    return [full[0]] + [row for row in full[1:] if row[0] in keep]


def build_investment_highlights(
    summary: FinancialSummary,
    industry: IndustryResearch,
    valuation: ValuationSummary,
) -> list[str]:
    """Three bullet points for the report cover (sell-side style)."""
    bullets: list[str] = []
    rating = recommendation_short(valuation)
    target = valuation.target_price
    price = summary.current_price
    if target and price and price > 0:
        pct = (target / price - 1) * 100
        blend_note = "analyst-led" if valuation.blend_weights.get("Analyst") == 1 else "multi-method"
        bullets.append(
            f"<b>{rating}</b> — blended target ${target:.2f} ({pct:+.1f}% vs ${price:.2f}); "
            f"{blend_note} valuation after outlier checks."
        )
    else:
        bullets.append(f"<b>{rating}</b> — blended target pending; see valuation cross-check on page 2.")

    health = build_financial_health_summary(summary)
    if health:
        bullets.append(health.rstrip("."))
    elif industry.key_trends:
        bullets.append(industry.key_trends[0])

    if not valuation.dcf_reliable and valuation.dcf_target:
        bullets.append(
            "Simon FCFE DCF not used in target (low FCFE yield / growth equity); "
            "rely on comps and analyst consensus where shown."
        )
    elif industry.key_trends and len(bullets) < 3:
        bullets.append(industry.key_trends[0])
    elif industry.risks and len(bullets) < 3:
        bullets.append(f"Key risk: {industry.risks[0]}")

    if industry.risks and len(bullets) < 3:
        bullets.append(f"Monitor: {industry.risks[0]}")

    return bullets[:3]


def build_executive_summary_pdf(
    summary: FinancialSummary,
    industry: IndustryResearch,
    valuation: ValuationSummary,
) -> str:
    name = summary.company_name
    rating = recommendation_short(valuation)
    target = f"${valuation.target_price:.2f}" if valuation.target_price else "N/A"
    overview = _truncate(industry.business_summary, 160)
    return (
        f"{name} ({summary.ticker}) — {summary.sector}. {overview} "
        f"<b>{rating}</b> rating; blended target {target}. "
        f"See Investment Highlights above and valuation cross-check on page 2."
    )


def _pdf_page_decorator(meta: dict[str, str], total_pages: int = 2):
    def _draw(canvas, doc) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica", 6)
        canvas.setFillColor(colors.HexColor("#4a5568"))
        left = doc.leftMargin
        right = doc.pagesize[0] - doc.rightMargin
        y_footer = 0.22 * inch
        header = (
            f"{meta['ticker']} | {meta['rating']} | PT {meta['target']} | "
            f"Data as of {meta['data_as_of']} | Latest: {meta['fiscal_period']}"
        )
        canvas.drawString(left, doc.pagesize[1] - 0.2 * inch, header)
        canvas.drawRightString(right, y_footer, f"Page {doc.page} of {total_pages}")
        canvas.drawString(
            left, y_footer,
            "Source: Yahoo Finance · Model output · Not investment advice",
        )
        canvas.restoreState()

    return _draw


def generate_pdf_report(
    summary: FinancialSummary,
    industry: IndustryResearch,
    competitors: CompetitorAnalysis,
    valuation: ValuationSummary,
    history,
    output_path: str,
    data: CompanyData | None = None,
) -> str:
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    fiscal_period = get_financial_period_key(data) if data else "N/A"
    data_as_of = datetime.now().strftime("%Y-%m-%d")
    pdf_meta = {
        "ticker": summary.ticker,
        "rating": recommendation_short(valuation),
        "target": f"${valuation.target_price:.2f}" if valuation.target_price else "N/A",
        "data_as_of": data_as_of,
        "fiscal_period": fiscal_period or "N/A",
    }
    page_draw = _pdf_page_decorator(pdf_meta, total_pages=2)

    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        leftMargin=0.38 * inch,
        rightMargin=0.38 * inch,
        topMargin=0.42 * inch,
        bottomMargin=0.38 * inch,
    )

    styles = getSampleStyleSheet()
    report_label = ParagraphStyle(
        "ReportLabel", parent=styles["Normal"], fontSize=7, textColor=ACCENT,
        fontName="Helvetica-Bold", spaceAfter=0,
    )
    title_style = ParagraphStyle(
        "Title", parent=styles["Heading1"], fontSize=13, textColor=NAVY,
        spaceAfter=1, alignment=TA_LEFT, fontName="Helvetica-Bold",
    )
    subtitle_style = ParagraphStyle("Subtitle", fontSize=7, textColor=colors.grey, spaceAfter=2)
    elements: list = []

    rating = recommendation_short(valuation)
    rec_color = GREEN if rating in ("BUY", "OVERWEIGHT") else (
        RED if rating in ("SELL", "UNDERWEIGHT") else colors.HexColor("#744210")
    )
    report_date = datetime.now().strftime("%B %d, %Y")
    price_str = f"${summary.current_price:.2f}" if summary.current_price else "N/A"
    target_str = f"${valuation.target_price:.2f}" if valuation.target_price else "N/A"
    upside_hdr = ""
    if valuation.target_price and summary.current_price and summary.current_price > 0:
        up = (valuation.target_price / summary.current_price - 1) * 100
        upside_hdr = f"<br/><font size='7'>Implied upside: {up:+.1f}%</font>"

    elements.append(Paragraph("EQUITY RESEARCH · COMPANY DEEP DIVE", report_label))
    header_data = [
        [
            Paragraph(f"{summary.company_name} ({summary.ticker}.US)", title_style),
            Paragraph(
                f'<font color="#{rec_color.hexval()[2:]}"><b>{rating}</b></font><br/>'
                f"<font size='8'>Blended Target: {target_str}</font>"
                f"<br/><font size='6'>DCF + Comps + Analyst (see cross-check table)</font>"
                f"{upside_hdr}",
                ParagraphStyle("RecBox", fontSize=11, alignment=TA_CENTER, textColor=rec_color),
            ),
        ],
        [
            Paragraph(
                f"{summary.sector} · {summary.industry}<br/>"
                f"{industry.market_cap_category} · Report Date: {report_date}<br/>"
                f"Latest financials: {fiscal_period} · Model data as of {data_as_of}",
                subtitle_style,
            ),
            Paragraph(
                f"<b>Price ({report_date.split(',')[0][:3]}):</b> {price_str}",
                ParagraphStyle("PriceBox", fontSize=8, alignment=TA_CENTER, textColor=NAVY),
            ),
        ],
    ]
    header_table = Table(header_data, colWidths=[4.7 * inch, 2.5 * inch])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (1, 0), (1, 0), 1, rec_color),
        ("BACKGROUND", (1, 0), (1, 1), colors.HexColor("#f8fafc")),
        ("LINEBELOW", (0, -1), (-1, -1), 1.5, NAVY),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (1, 0), (1, -1), 6),
        ("RIGHTPADDING", (1, 0), (1, -1), 6),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 2))

    elements.append(_section_title("Investment Highlights"))
    for bullet in build_investment_highlights(summary, industry, valuation):
        elements.append(_body(f"• {bullet}", 5.5))
    elements.append(Spacer(1, 2))

    elements.append(_section_title("Investment Thesis / Executive Summary"))
    elements.append(_body(build_executive_summary_pdf(summary, industry, valuation), 5.3))
    elements.append(Spacer(1, 1.5))

    shares = get_shares_outstanding(data.info) if data and data.info else None
    pb = get_info_value(data.info, "priceToBook") if data else None
    avg_vol = get_info_value(data.info, "averageVolume") if data else None

    key_data_rows = [
        [_th_cell("Item"), _th_cell("Value"), _th_cell("Item"), _th_cell("Value")],
        [_cell("Market Cap"), _cell(format_large_number(summary.market_cap)),
         _cell("Shares Out."), _cell(f"{shares / 1e6:.1f}M" if shares else "N/A")],
        [_cell("52-Week Range"), _cell(summary.fifty_two_week_range),
         _cell("P/E (TTM)"), _cell(_fmt_ratio(summary.pe_ratio))],
        [_cell("P/B"), _cell(_fmt_ratio(pb) if pb else "N/A"),
         _cell("Beta"), _cell(f"{summary.beta:.2f}" if summary.beta else "N/A")],
        [_cell("Avg Volume"), _cell(f"{avg_vol / 1e6:.2f}M" if avg_vol else "N/A"),
         _cell("Analyst Rec."), _cell(summary.recommendation_key.title())],
        [_cell("Revenue (TTM)"), _cell(format_large_number(summary.revenue)),
         _cell("Rev Growth"), _cell(_fmt_pct(summary.revenue_growth, signed=True))],
        [_cell("Op. Margin"), _cell(_fmt_pct(summary.operating_margin)),
         _cell("FCF (TTM)"), _cell(format_large_number(summary.free_cash_flow))],
    ]
    key_table = Table(key_data_rows, colWidths=[1.0 * inch, 1.15 * inch, 1.0 * inch, 1.15 * inch])

    chart_buf = _price_chart(history, summary.ticker)
    if chart_buf:
        top_row = Table(
            [[key_table, Image(chart_buf, width=2.75 * inch, height=0.95 * inch)]],
            colWidths=[4.15 * inch, 2.9 * inch],
        )
    else:
        top_row = key_table
    top_row.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    key_table.setStyle(_table_style_header())
    elements.append(_section_title("Key Data & Market Highlights"))
    elements.append(top_row)

    if data:
        fv_rows = build_financial_valuation_summary_table(data, summary)
        if fv_rows:
            elements.append(Spacer(1, 1.5))
            elements.append(_section_title("Financial & Valuation Summary"))
            elements.append(_body("(USD millions except per-share; E = forecast)", 5))
            ncols = len(fv_rows[0])
            metric_w = 1.38 * inch
            year_w = (7.24 * inch - metric_w) / max(ncols - 1, 1)
            elements.append(_make_styled_table(fv_rows, [metric_w] + [year_w] * (ncols - 1)))

    elements.append(PageBreak())

    # --- Page 2 only: valuation cross-check + industry + peers (strict 2-page layout) ---
    page2_hdr = ParagraphStyle(
        "Page2Hdr", fontName="Helvetica-Bold", fontSize=9, textColor=NAVY, spaceAfter=2,
    )
    elements.append(Paragraph(
        f"<b>Page 2 of 2</b> · {summary.company_name} ({summary.ticker}) · "
        f"Valuation · Industry · Peers · Data as of {data_as_of} · {fiscal_period}",
        page2_hdr,
    ))

    cross_rows = build_valuation_crosscheck_table(valuation, summary)
    elements.append(Spacer(1, 2))
    elements.append(_section_title("Valuation Cross-Check"))
    elements.append(_make_styled_table(
        cross_rows,
        [1.1 * inch, 0.72 * inch, 0.62 * inch, 0.78 * inch, 2.72 * inch],
        wrap_col=4,
    ))
    elements.append(Spacer(1, 2))
    elements.append(_section_title("DCF Valuation Summary (Simon FCFE)"))
    elements.append(_make_styled_table(
        build_dcf_detail_table_compact(valuation.dcf),
        [1.55 * inch, 2.1 * inch],
    ))
    elements.append(Spacer(1, 4))

    industry_overview = _truncate(industry.business_summary, 420)
    industry_text = (
        f"<b>1. Industry Overview</b><br/>{industry_overview}<br/><br/>"
        f"<b>2. Key Trends</b><br/>" + "<br/>".join(f"• {t}" for t in industry.key_trends) + "<br/><br/>"
        f"<b>3. Risk Factors</b><br/>" + "<br/>".join(f"• {r}" for r in industry.risks)
    )

    comp_rows = [[
        _th_cell("Ticker"), _th_cell("Company"), _th_cell("Market Cap"),
        _th_cell("P/E"), _th_cell("EV/EBITDA"), _th_cell("Net Margin"), _th_cell("Rev Growth"),
    ]]
    for p in competitors.peers:
        comp_rows.append([
            _cell(p.ticker), _cell(p.name[:22]), _cell(format_large_number(p.market_cap)),
            _cell(_fmt_ratio(p.pe_ratio)), _cell(_fmt_ratio(p.ev_ebitda)),
            _cell(_fmt_pct(p.profit_margin)), _cell(_fmt_pct(p.revenue_growth, signed=True)),
        ])
    comp_rows.append([
        _cell(f"{summary.ticker}*"), _cell(summary.company_name[:22]),
        _cell(format_large_number(summary.market_cap)), _cell(_fmt_ratio(summary.pe_ratio)),
        _cell(_fmt_ratio(summary.ev_ebitda)), _cell(_fmt_pct(summary.profit_margin)),
        _cell(_fmt_pct(summary.revenue_growth, signed=True)),
    ])
    comp_table = Table(
        comp_rows,
        colWidths=[0.55 * inch, 1.35 * inch, 0.85 * inch, 0.55 * inch, 0.75 * inch, 0.7 * inch, 0.7 * inch],
    )
    sty = _table_style_header()
    sty.add("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#ebf8ff"))
    comp_table.setStyle(sty)

    elements.append(_section_title("Industry Analysis"))
    elements.append(_body(industry_text, 5.5))
    elements.append(Spacer(1, 4))
    elements.append(_section_title("Competitor Analysis"))
    comp_table.splitByRow = 0
    elements.append(comp_table)
    elements.append(Spacer(1, 1))
    elements.append(_body(
        f"* Subject company. {competitors.relative_position}. "
        f"{build_financial_health_summary(summary)} "
        f"Source: Yahoo Finance. Not investment advice.",
        5,
    ))

    doc.build(elements, onFirstPage=page_draw, onLaterPages=page_draw)
    return output_path


def generate_summary_pdf(rows: list[ReportRow], output_path: str) -> str:
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    page = landscape(letter)
    doc = SimpleDocTemplate(
        output_path,
        pagesize=page,
        leftMargin=0.3 * inch,
        rightMargin=0.3 * inch,
        topMargin=0.3 * inch,
        bottomMargin=0.3 * inch,
    )
    elements: list = []
    title_style = ParagraphStyle(
        "SumTitle", fontName="Helvetica-Bold", fontSize=13, textColor=NAVY, spaceAfter=3,
    )
    elements.append(Paragraph(
        f"US Top {len(rows)} by Market Cap — Equity Research Summary",
        title_style,
    ))
    elements.append(Paragraph(
        f"Report Date: {datetime.now().strftime('%B %d, %Y')} · "
        "Blended target excludes outlier methods (DCF/Comps/Analyst cross-check)",
        ParagraphStyle("Sub", fontSize=7.5, textColor=colors.grey),
    ))
    elements.append(Spacer(1, 4))

    header = [
        "#", "Ticker", "Company", "Sector", "Price", "Mkt Cap", "P/E",
        "Rev%", "OpMgn%", "DCF", "Comps", "Target", "Ups%", "Rating",
    ]
    table_data = [[_th_cell(h) for h in header]]
    for rank, r in enumerate(rows, start=1):
        table_data.append([
            _cell(str(rank)),
            _cell(r.ticker),
            _cell(r.company_name[:18]),
            _cell(r.sector[:14]),
            _cell(f"${r.current_price:.2f}" if r.current_price else "N/A"),
            _cell(format_large_number(r.market_cap)),
            _cell(_fmt_ratio(r.pe_ratio)),
            _cell(_fmt_pct(r.revenue_growth, signed=True)),
            _cell(_fmt_pct(r.operating_margin)),
            _cell(f"${r.dcf_fair_value:.0f}" if r.dcf_fair_value else "N/A"),
            _cell(f"${r.comps_fair_value:.0f}" if r.comps_fair_value else "N/A"),
            _cell(f"${r.target_price:.2f}" if r.target_price else "N/A"),
            _cell(_fmt_pct(r.upside_pct, signed=True)),
            _cell(r.rating),
        ])

    usable = page[0] - 0.6 * inch
    col_widths = [
        0.22, 0.35, 0.85, 0.7, 0.42, 0.5, 0.32, 0.35, 0.35, 0.42, 0.42, 0.48, 0.42, 0.42,
    ]
    scale = usable / sum(col_widths)
    col_widths = [w * scale for w in col_widths]

    summary_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    summary_table.splitByRow = 0
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 5.2),
        ("GRID", (0, 0), (-1, -1), 0.2, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_BG]),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 6))

    buy_count = sum(1 for r in rows if r.rating in ("BUY", "OVERWEIGHT"))
    sell_count = sum(1 for r in rows if r.rating in ("SELL", "UNDERWEIGHT"))
    hold_count = len(rows) - buy_count - sell_count
    elements.append(_body(
        f"<b>Aggregate view:</b> {buy_count} Buy/Overweight, {hold_count} Hold, {sell_count} Sell/Underweight "
        f"among the top {len(rows)} US companies by market capitalization. "
        f"DCF and comparable multiples are model-based estimates. Not investment advice.",
        6,
    ))
    doc.build(elements)

    csv_path = output_path.replace(".pdf", ".csv")
    try:
        pd.DataFrame([{
            "Rank": i,
            "Ticker": r.ticker,
            "Company": r.company_name,
            "Sector": r.sector,
            "Price": r.current_price,
            "Market_Cap": r.market_cap,
            "PE": r.pe_ratio,
            "Revenue_Growth_Pct": r.revenue_growth,
            "Operating_Margin_Pct": r.operating_margin,
            "DCF_Fair_Value": r.dcf_fair_value,
            "Comps_Fair_Value": r.comps_fair_value,
            "Target_Price": r.target_price,
            "Upside_Pct": r.upside_pct,
            "Rating": r.rating,
            "Financial_Health": r.financial_health,
        } for i, r in enumerate(rows, start=1)]).to_csv(csv_path, index=False)
        print(f"Summary CSV saved: {csv_path}")
    except OSError as e:
        alt = csv_path.replace(".csv", f"_{datetime.now().strftime('%H%M%S')}.csv")
        pd.DataFrame([{"Ticker": r.ticker} for r in rows]).to_csv(alt, index=False)
        print(f"Summary CSV locked; saved to {alt} ({e})", file=sys.stderr)
    return output_path


# ---------------------------------------------------------------------------
# Simon DCF Excel export
# ---------------------------------------------------------------------------


def _historical_fcfe_detail(data: CompanyData, max_years: int = 4) -> list[dict[str, Any]]:
    cf = data.cashflow
    bs = data.balance_sheet
    if cf is None or cf.empty:
        return []
    n = min(max_years, cf.shape[1])
    details: list[dict[str, Any]] = []
    for col_i in range(n - 1, -1, -1):
        try:
            dt = cf.columns[col_i]
            year = dt.year if hasattr(dt, "year") else str(dt)[:4]
        except Exception:
            year = f"Y{col_i}"
        ocf = _safe_series_value(cf, "Operating Cash Flow", col_i)
        inv_cf = _safe_series_value(cf, "Investing Cash Flow", col_i)
        if inv_cf is None:
            capex = _safe_series_value(cf, "Capital Expenditure", col_i)
            inv_cf = capex if capex is not None else 0.0
        debt_chg = 0.0
        if bs is not None and not bs.empty:
            d0 = _safe_series_value(bs, "Total Debt", col_i)
            d1 = _safe_series_value(bs, "Total Debt", col_i + 1)
            if d0 is not None and d1 is not None:
                debt_chg = d0 - d1
        cash_chg = _safe_series_value(cf, "Changes In Cash", col_i) or 0.0
        inc_cash = max(0.0, -cash_chg) if cash_chg else 0.0
        simon_fcfe = (ocf or 0) + (inv_cf or 0) + debt_chg
        standard = _standard_fcf_for_year(cf, col_i)
        used = standard if standard and standard > 0 and simon_fcfe < standard * 0.5 else simon_fcfe
        details.append({
            "year": year,
            "ocf": ocf,
            "investing_cf": inv_cf,
            "debt_increase": debt_chg,
            "increase_operating_cash": inc_cash,
            "simon_fcfe": simon_fcfe,
            "yahoo_fcf": standard,
            "fcfe_used": used,
        })
    return details


def _fetch_peer_ps(ticker: str) -> float | None:
    try:
        ps = get_info_value(yf.Ticker(ticker).info or {}, "priceToSalesTrailing12Months")
        return float(ps) if ps else None
    except Exception:
        return None


def _excel_style_header(cell, hdr_fill, hdr_font) -> None:
    cell.fill = hdr_fill
    cell.font = hdr_font


def _add_peer_multiples_sheet(
    wb,
    comp_analysis: CompetitorAnalysis,
    summary: FinancialSummary,
    comps: CompsResult,
    hdr_fill,
    hdr_font,
    title_font,
    num_fmt_px,
) -> dict[str, str]:
    """Peer table with MEDIAN formulas; returns cell refs for comps sheet."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Peer Multiples")
    ws["A1"] = "Comparable company multiples (Yahoo Finance)"
    ws["A1"].font = title_font
    ws["A2"] = "P/E and EV/EBITDA medians include subject; P/S median uses peers only (matches Python model)."

    headers = ["Ticker", "Company", "P/E", "EV/EBITDA", "P/S"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        _excel_style_header(cell, hdr_fill, hdr_font)

    row = 5
    peer_ps: list[float] = []
    for p in comp_analysis.peers:
        ps = _fetch_peer_ps(p.ticker)
        if ps:
            peer_ps.append(ps)
        ws.cell(row=row, column=1, value=p.ticker)
        ws.cell(row=row, column=2, value=p.name)
        ws.cell(row=row, column=3, value=p.pe_ratio)
        ws.cell(row=row, column=4, value=p.ev_ebitda)
        ws.cell(row=row, column=5, value=ps)
        for col in (3, 4, 5):
            ws.cell(row=row, column=col).number_format = "0.00x" if col < 5 else "0.00x"
        row += 1

    peer_start = 5
    peer_end = row - 1
    sub_row = row
    ws.cell(row=sub_row, column=1, value=comp_analysis.target_ticker)
    ws.cell(row=sub_row, column=2, value=summary.company_name[:24])
    ws.cell(row=sub_row, column=3, value=summary.pe_ratio)
    ws.cell(row=sub_row, column=4, value=summary.ev_ebitda)
    sub_ps = _fetch_peer_ps(comp_analysis.target_ticker)
    ws.cell(row=sub_row, column=5, value=sub_ps)
    for col in range(1, 6):
        ws.cell(row=sub_row, column=col).font = Font(bold=True)

    med_row = sub_row + 1
    ws.cell(row=med_row, column=2, value="Median P/E (peers + subject)")
    ws.cell(row=med_row, column=3, value=f"=MEDIAN(C{peer_start}:C{sub_row})")
    med_ev_row = med_row + 1
    ws.cell(row=med_ev_row, column=2, value="Median EV/EBITDA (peers + subject)")
    ws.cell(row=med_ev_row, column=4, value=f"=MEDIAN(D{peer_start}:D{sub_row})")
    med_ps_row = med_ev_row + 1
    ws.cell(row=med_ps_row, column=2, value="Median P/S (peers only)")
    if peer_end >= peer_start:
        ws.cell(row=med_ps_row, column=5, value=f"=MEDIAN(E{peer_start}:E{peer_end})")
    else:
        ws.cell(row=med_ps_row, column=5, value=comps.peer_median_ps)

    ws.cell(row=med_row + 4, column=1, value="Python reference (peer_median_pe)")
    ws.cell(row=med_row + 4, column=3, value=comps.peer_median_pe)
    ws.cell(row=med_row + 5, column=1, value="Python reference (peer_median_ev_ebitda)")
    ws.cell(row=med_row + 5, column=4, value=comps.peer_median_ev_ebitda)
    ws.cell(row=med_row + 6, column=1, value="Python reference (peer_median_ps)")
    ws.cell(row=med_row + 6, column=5, value=comps.peer_median_ps)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    for col in ("C", "D", "E"):
        ws.column_dimensions[col].width = 14

    return {
        "med_pe": f"'Peer Multiples'!C{med_row}",
        "med_ev": f"'Peer Multiples'!D{med_ev_row}",
        "med_ps": f"'Peer Multiples'!E{med_ps_row}",
    }


def _add_comps_valuation_sheet(
    wb,
    data: CompanyData,
    summary: FinancialSummary,
    comps: CompsResult,
    peer_refs: dict[str, str],
    hdr_fill,
    hdr_font,
    title_font,
    num_fmt_m,
    num_fmt_px,
) -> dict[str, str]:
    """Comps implied prices with Excel formulas."""
    ws = wb.create_sheet("Comps Valuation")
    ws["A1"] = "Comparable multiples — implied share price"
    ws["A1"].font = title_font
    ws["A2"] = "Implied price = f(peer median multiple, subject fundamentals, shares outstanding)."

    info = data.info
    eps = get_info_value(info, "trailingEps")
    ebitda = get_ebitda(data.financials, info)
    revenue = get_revenue(data.financials, info)
    net_debt = get_net_debt(data.balance_sheet) or 0
    shares = get_shares_outstanding(info)

    fund = [
        ("Trailing EPS ($/sh)", eps),
        ("EBITDA ($)", ebitda),
        ("Revenue ($)", revenue),
        ("Net debt ($)", net_debt),
        ("Shares outstanding", shares),
        ("Current price ($/sh)", summary.current_price),
    ]
    ws["A4"] = "Subject fundamentals"
    ws["A4"].font = title_font
    for i, (lab, val) in enumerate(fund, start=5):
        ws.cell(row=i, column=1, value=lab)
        ws.cell(row=i, column=2, value=val)
        fmt = num_fmt_px if "EPS" in lab or "price" in lab else num_fmt_m
        if "Shares" in lab:
            fmt = "#,##0"
        ws.cell(row=i, column=2).number_format = fmt

    r_eps, r_ebitda, r_rev, r_nd, r_sh, r_px = 5, 6, 7, 8, 9, 10

    ws["A12"] = "Method"
    ws["B12"] = "Calculation (formula)"
    ws["C12"] = "Implied $/sh"
    for c in range(1, 4):
        _excel_style_header(ws.cell(row=12, column=c), hdr_fill, hdr_font)

    ws["A13"] = "P/E (peer median × EPS)"
    ws["B13"] = f"Median P/E × B{r_eps} (EPS)"
    ws["C13"] = (
        f"=IF(OR({peer_refs['med_pe']}=\"\",B{r_eps}=\"\",B{r_eps}<=0),\"\","
        f"{peer_refs['med_pe']}*B{r_eps})"
    )

    ws["A14"] = "EV/EBITDA (implied equity / shares)"
    ws["B14"] = f"(Median EV/EBITDA × B{r_ebitda} − B{r_nd}) / B{r_sh}"
    ws["C14"] = (
        f"=IF(OR({peer_refs['med_ev']}=\"\",B{r_ebitda}=\"\",B{r_ebitda}<=0,B{r_sh}=\"\",B{r_sh}<=0),"
        f"\"\",(({peer_refs['med_ev']}*B{r_ebitda})-B{r_nd})/B{r_sh})"
    )

    ws["A15"] = "P/S (implied equity / shares)"
    ws["B15"] = f"(Median P/S × B{r_rev}) / B{r_sh}"
    ws["C15"] = (
        f"=IF(OR({peer_refs['med_ps']}=\"\",B{r_rev}=\"\",B{r_rev}<=0,B{r_sh}=\"\",B{r_sh}<=0),"
        f"\"\",({peer_refs['med_ps']}*B{r_rev})/B{r_sh})"
    )

    ws["A16"] = "Comps blended (average of available)"
    ws["B16"] = "=AVERAGE(C13:C15)"
    ws["C16"] = "=IF(COUNT(C13,C14,C15)=0,\"\",AVERAGE(C13,C14,C15))"

    for r in (13, 14, 15, 16):
        ws.cell(row=r, column=3).number_format = num_fmt_px

    ws.cell(row=18, column=1, value="Python reference (blended_fair_value)")
    ws.cell(row=18, column=2, value=comps.blended_fair_value)
    ws.cell(row=18, column=2).number_format = num_fmt_px

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 16

    return {"comps_blended": "C16", "current": f"B{r_px}"}


def _add_analyst_sheet(
    wb,
    summary: FinancialSummary,
    hdr_fill,
    hdr_font,
    title_font,
    num_fmt_px,
    num_fmt_pct,
) -> dict[str, str]:
    ws = wb.create_sheet("Analyst Consensus")
    ws["A1"] = "Analyst consensus (Yahoo Finance targetMeanPrice)"
    ws["A1"].font = title_font
    ws["A2"] = "Used as one input to the blended target (no formula derivation — market consensus)."

    ws["A4"] = "Field"
    ws["B4"] = "Value"
    _excel_style_header(ws["A4"], hdr_fill, hdr_font)
    _excel_style_header(ws["B4"], hdr_fill, hdr_font)

    ws["A5"] = "Analyst mean target ($/sh)"
    ws["B5"] = summary.analyst_target
    ws["B5"].number_format = num_fmt_px
    ws["A6"] = "Current price ($/sh)"
    ws["B6"] = summary.current_price
    ws["B6"].number_format = num_fmt_px
    ws["A7"] = "Upside vs. current"
    ws["B7"] = '=IF(OR(B5="",B6="",B6=0),"",B5/B6-1)'
    ws["B7"].number_format = num_fmt_pct

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    return {"analyst_target": "B5", "current": "B6"}


def _add_valuation_summary_sheet(
    wb,
    valuation: ValuationSummary,
    comps_refs: dict[str, str],
    analyst_refs: dict[str, str],
    hdr_fill,
    hdr_font,
    title_font,
    num_fmt_px,
    num_fmt_pct,
) -> None:
    ws = wb.create_sheet("Valuation Summary")
    ws["A1"] = "Multi-method valuation — blended target"
    ws["A1"].font = title_font
    ws["A2"] = (
        "Base weights: DCF 28%, Comps 44%, Analyst 28%. "
        "Effective weights reflect outlier / reliability exclusions from the Python model."
    )

    headers = ["Method", "Implied $/sh", "vs. Current", "Base weight", "Effective weight", "Notes"]
    for c, h in enumerate(headers, start=1):
        _excel_style_header(ws.cell(row=4, column=c, value=h), hdr_fill, hdr_font)

    current = f"'Comps Valuation'!{comps_refs['current']}"
    base_w = {"DCF": 0.28, "Comps": 0.44, "Analyst": 0.28}

    methods = [
        (
            "DCF (Simon FCFE)",
            "=DCF!B28",
            "DCF",
            valuation.excluded_methods.get("DCF", "Included in blend" if valuation.dcf_reliable else ""),
        ),
        (
            "Comps (blended multiples)",
            f"='Comps Valuation'!{comps_refs['comps_blended']}",
            "Comps",
            valuation.excluded_methods.get("Comps", "Included in blend" if valuation.comps_reliable else ""),
        ),
        (
            "Analyst consensus",
            f"='Analyst Consensus'!{analyst_refs['analyst_target']}",
            "Analyst",
            valuation.excluded_methods.get("Analyst", "Included in blend" if valuation.analyst_reliable else ""),
        ),
    ]

    for i, (label, price_formula, key, note) in enumerate(methods, start=5):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=price_formula)
        ws.cell(row=i, column=2).number_format = num_fmt_px
        ws.cell(row=i, column=3, value=f"=IF(OR(B{i}=\"\",{current}=\"\",{current}=0),\"\",B{i}/{current}-1)")
        ws.cell(row=i, column=3).number_format = num_fmt_pct
        ws.cell(row=i, column=4, value=base_w[key])
        ws.cell(row=i, column=4).number_format = num_fmt_pct
        eff = valuation.blend_weights.get(key, 0.0)
        ws.cell(row=i, column=5, value=eff if key in valuation.blend_weights else 0.0)
        ws.cell(row=i, column=5).number_format = num_fmt_pct
        ws.cell(row=i, column=6, value=note or ("Included in blend" if eff else "Excluded"))

    blend_row = 9
    ws.cell(row=blend_row, column=1, value="Blended target (effective weights)")
    ws.cell(row=blend_row, column=2, value="=SUMPRODUCT(E5:E7,B5:B7)")
    ws.cell(row=blend_row, column=2).number_format = num_fmt_px
    ws.cell(row=blend_row, column=3, value=f"=IF(OR(B{blend_row}=\"\",{current}=\"\",{current}=0),\"\",B{blend_row}/{current}-1)")
    ws.cell(row=blend_row, column=3).number_format = num_fmt_pct
    ws.cell(row=blend_row, column=4, value=1.0)
    ws.cell(row=blend_row, column=4).number_format = num_fmt_pct
    ws.cell(row=blend_row, column=6, value=valuation.recommendation)

    ws.cell(row=blend_row + 2, column=1, value="Formula check (if all methods included, normalized base weights)")
    ws.cell(row=blend_row + 2, column=2, value="=SUMPRODUCT($D$5:$D$7/SUM($D$5:$D$7),B5:B7)")
    ws.cell(row=blend_row + 2, column=2).number_format = num_fmt_px

    ws.cell(row=blend_row + 3, column=1, value="Python reference (target_price)")
    ws.cell(row=blend_row + 3, column=2, value=valuation.target_price)
    ws.cell(row=blend_row + 3, column=2).number_format = num_fmt_px

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 44


def export_simon_valuation_excel(
    ticker: str,
    data: CompanyData,
    summary: FinancialSummary,
    valuation: ValuationSummary,
    comp_analysis: CompetitorAnalysis,
    output_path: str,
) -> str:
    """
    Export Simon-style FCFE/DCF workbook with comps, analyst, and blended summary (formula-linked).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    dcf = valuation.dcf
    comps = valuation.comps

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1A365D")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    title_font = Font(bold=True, size=12, color="1A365D")
    num_fmt_m = "#,##0"
    num_fmt_pct = "0.00%"
    num_fmt_px = "$#,##0.00"

    info = data.info
    beta = get_info_value(info, "beta") or 1.0
    rf, erp = 0.04, 0.055
    ke = dcf.wacc
    hist = _historical_fcfe_detail(data)
    years_proj = dcf.projection_years or DEFAULT_PROJECTION_YEARS

    # --- Sheet 1: Assumptions ---
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a["A1"] = "Simon FCFE DCF — Assumptions"
    ws_a["A1"].font = title_font
    labels = [
        ("Company", summary.company_name),
        ("Ticker", ticker.upper()),
        ("Valuation date", datetime.now().strftime("%Y-%m-%d")),
        ("", ""),
        ("CAPM — Risk-free rate (Rf)", rf),
        ("CAPM — Equity risk premium (ERP)", erp),
        ("Beta", beta),
        ("Cost of common equity (Ke)", ke),
        ("Terminal growth (g)", dcf.terminal_growth),
        ("FCFE growth (forecast)", dcf.fcf_growth),
        ("Projection years", years_proj),
        ("Base FCFE (avg. historical, $)", dcf.fcf_base),
        ("Shares outstanding", dcf.shares_used),
        ("Current share price ($)", summary.current_price),
        ("Implied price — DCF ($)", dcf.implied_price),
    ]
    for i, (lab, val) in enumerate(labels, start=3):
        ws_a[f"A{i}"] = lab
        if lab == "Cost of common equity (Ke)":
            ws_a[f"B{i}"] = "=B7+B9*B8"
        else:
            ws_a[f"B{i}"] = val
        if lab.endswith("(Rf)") or lab.endswith("(ERP)") or lab in ("Cost of common equity (Ke)", "Terminal growth (g)", "FCFE growth (forecast)"):
            ws_a[f"B{i}"].number_format = num_fmt_pct
        if "FCFE" in lab or "price" in lab.lower():
            if "growth" not in lab.lower() and "years" not in lab.lower():
                ws_a[f"B{i}"].number_format = num_fmt_m if "Shares" not in lab else "#,##0"
        if "price" in lab.lower():
            ws_a[f"B{i}"].number_format = num_fmt_px

    ke_row, g_row, growth_row, base_row, shares_row = 10, 11, 12, 14, 15

    # --- Sheet 2: Historical FCFE build ---
    ws_h = wb.create_sheet("Historical FCFE")
    ws_h["A1"] = "Historical FCFE Build (Simon components)"
    ws_h["A1"].font = title_font
    ws_h["A2"] = "FCFE = Cash from operations + Cash from investing + Increase in debt"
    h_headers = ["Line item ($)"] + [str(h["year"]) for h in hist]
    for c, h in enumerate(h_headers, start=1):
        cell = ws_h.cell(row=4, column=c, value=h)
        cell.fill, cell.font = hdr_fill, hdr_font
    h_rows = [
        ("+ Cash from operations", "ocf"),
        ("+ Cash from investing", "investing_cf"),
        ("+ Increase in debt", "debt_increase"),
        ("- Increase in operating cash (MAX(0,-ΔCash))", "increase_operating_cash"),
        ("= Simon FCFE (sum)", "simon_fcfe"),
        ("Yahoo Free Cash Flow (reference)", "yahoo_fcf"),
        ("FCFE used in valuation model", "fcfe_used"),
    ]
    for ri, (label, key) in enumerate(h_rows, start=5):
        ws_h.cell(row=ri, column=1, value=label)
        for ci, h in enumerate(hist, start=2):
            v = h.get(key)
            cell = ws_h.cell(row=ri, column=ci, value=v)
            cell.number_format = num_fmt_m
    avg_row = 5 + len(h_rows)
    ws_h.cell(row=avg_row, column=1, value="Average FCFE used as base")
    ws_h.cell(row=avg_row, column=2, value=dcf.fcf_base)
    ws_h.cell(row=avg_row, column=2).number_format = num_fmt_m

    # --- Sheet 3: DCF (mirror Simon DCF sheet) ---
    ws_d = wb.create_sheet("DCF")
    ws_d["A1"] = "DCF Valuation"
    ws_d["A1"].font = title_font
    ws_d["A2"] = "Date of Valuation"
    ws_d["B2"] = datetime.now().strftime("%Y-%m-%d")
    ws_d["A3"] = "Cost of Common Equity (Ke)"
    ws_d["B3"] = f"=Assumptions!B{ke_row}"
    ws_d["B3"].number_format = num_fmt_pct

    start_col = 2
    last_fc_col = start_col + years_proj - 1
    term_col = last_fc_col + 1

    ws_d.cell(row=5, column=term_col, value="Terminal year")
    for i in range(years_proj):
        col = start_col + i
        yr = datetime.now().year + i + 1
        ws_d.cell(row=5, column=col, value=f"FY {yr}E")
        cletter = get_column_letter(col)
        ws_d.cell(row=16, column=col, value=yr)

    ws_d["A6"] = "Free Cash Flow to Common Equity (forecast)"
    ws_d["A17"] = "Valuation to Common Equity"
    ws_d["A18"] = "Free Cash Flow to Common Equity"
    ws_d["A19"] = "Present Value of FCF"

    base_cell = f"Assumptions!B{base_row}"
    growth_cell = f"Assumptions!B{growth_row}"
    for i in range(years_proj):
        col = start_col + i
        cletter = get_column_letter(col)
        if i == 0:
            ws_d.cell(row=18, column=col, value=f"={base_cell}*(1+{growth_cell})")
        else:
            prev = get_column_letter(col - 1)
            ws_d.cell(row=18, column=col, value=f"={prev}18*(1+{growth_cell})")
        ws_d.cell(row=19, column=col, value=f"={cletter}18/(1+$B$3)^{i + 1}")

    ws_d["A20"] = "Present Value Beyond 5 Years"
    g_cell = f"Assumptions!B{g_row}"
    term_l = get_column_letter(term_col)
    last_l = get_column_letter(last_fc_col)
    ws_d["B20"] = f"=({last_l}18/($B$3-{g_cell}))/(1+$B$3)^{years_proj}"

    ws_d["A21"] = "Present Value of First 5 Years"
    first_l = get_column_letter(start_col)
    last_l19 = get_column_letter(last_fc_col)
    ws_d["B21"] = f"=SUM({first_l}19:{last_l19}19)"
    ws_d["H21"] = "g="
    ws_d["I21"] = f"=Assumptions!B{g_row}"
    ws_d["I21"].number_format = num_fmt_pct

    ws_d["A22"] = "Forecast Equity Value Before Time Adj."
    ws_d["B22"] = "=SUM(B20:B21)"
    ws_d["A23"] = "Forecasted Value as of Valuation Date (mid-year adj.)"
    ws_d["B23"] = "=B22*(1+B3/2)"
    ws_d["A24"] = "Excess Cash on Valuation Date"
    ws_d["B24"] = 0
    ws_d["A25"] = "Less Value of Contingent Equity Claims"
    ws_d["B25"] = 0
    ws_d["A26"] = "Value Attributable to Common Equity"
    ws_d["B26"] = "=B23"
    ws_d["A27"] = "Common Shares Outstanding"
    ws_d["B27"] = f"=Assumptions!B{shares_row}"
    ws_d["A28"] = "Forecast Price / Share"
    ws_d["B28"] = "=B26/B27"
    ws_d["B28"].number_format = num_fmt_px

    ws_d.cell(row=18, column=term_col, value=f"={last_l}18")

    ws_d["H17"] = "Python model (reference)"
    ws_d["H17"].font = Font(bold=True, italic=True)
    ref_rows = [
        (21, dcf.pv_explicit, num_fmt_m),
        (20, dcf.pv_terminal, num_fmt_m),
        (22, dcf.equity_value_raw, num_fmt_m),
        (23, dcf.equity_value, num_fmt_m),
        (26, dcf.equity_value, num_fmt_m),
        (28, dcf.implied_price, num_fmt_px),
    ]
    for row, val, fmt in ref_rows:
        if val is not None:
            ws_d.cell(row=row, column=8, value=val)
            ws_d.cell(row=row, column=8).number_format = fmt
    if dcf.projected_fcfe:
        ws_d["H18"] = "Projected FCFE ($)"
        for i, f in enumerate(dcf.projected_fcfe):
            ws_d.cell(row=19 + i, column=8, value=f)
            ws_d.cell(row=19 + i, column=8).number_format = num_fmt_m

    for row in (18, 19, 20, 21, 22, 23, 26, 28):
        for col in range(start_col, term_col + 1):
            ws_d.cell(row=row, column=col).number_format = num_fmt_m if row != 28 else num_fmt_px
    ws_d["B28"].number_format = num_fmt_px

    # --- Sheet 4: DCF Detail (values from Python, audit) ---
    ws_v = wb.create_sheet("DCF Audit")
    ws_v["A1"] = "DCF audit trail (computed values)"
    ws_v["A1"].font = title_font
    audit = build_dcf_detail_table(dcf)
    for ri, row in enumerate(audit, start=3):
        ws_v.cell(row=ri, column=1, value=row[0] if row else "")
        if len(row) > 1:
            ws_v.cell(row=ri, column=2, value=row[1])

    peer_refs = _add_peer_multiples_sheet(
        wb, comp_analysis, summary, comps, hdr_fill, hdr_font, title_font, num_fmt_px,
    )
    comps_refs = _add_comps_valuation_sheet(
        wb, data, summary, comps, peer_refs, hdr_fill, hdr_font, title_font, num_fmt_m, num_fmt_px,
    )
    analyst_refs = _add_analyst_sheet(
        wb, summary, hdr_fill, hdr_font, title_font, num_fmt_px, num_fmt_pct,
    )
    _add_valuation_summary_sheet(
        wb, valuation, comps_refs, analyst_refs, hdr_fill, hdr_font, title_font, num_fmt_px, num_fmt_pct,
    )

    for ws in wb.worksheets:
        ws.column_dimensions["A"].width = 42
        for col in range(2, 12):
            ws.column_dimensions[get_column_letter(col)].width = 14

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate institutional-style equity research PDF reports (English).",
    )
    parser.add_argument("ticker", nargs="?", default=None, help="Single ticker (e.g., AAPL). Omit with --batch-top20.")
    parser.add_argument(
        "--batch-top20", action="store_true",
        help="Batch: PDFs for top 5 + TSLA/BX + Top 20 summary table",
    )
    parser.add_argument("--pdf-count", type=int, default=TOP_US_PDF_REPORT_COUNT, help="Number of PDF reports (default: 5)")
    parser.add_argument("--summary-count", type=int, default=TOP_US_SUMMARY_COUNT, help="Summary universe size (default: 20)")
    parser.add_argument("--weekly", action="store_true", help="Only run on Mondays (for scheduled jobs)")
    parser.add_argument(
        "--refresh-stale",
        action="store_true",
        help="Only refresh tickers with new financial period or older than stale threshold",
    )
    parser.add_argument("--stale-days", type=int, default=STALE_MAX_AGE_DAYS, help="Max age in days before refresh (default: 7)")
    parser.add_argument("--force", action="store_true", help="Force full refresh; bypass weekly/stale skips")
    parser.add_argument("--replace-old", action="store_true", default=True, help="Remove prior reports before batch run (default: on)")
    parser.add_argument("--no-replace-old", action="store_false", dest="replace_old", help="Keep dated report files")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="Output directory (default: reports/)")
    parser.add_argument("--peers", nargs="*", default=None, help="Optional competitor tickers for single-ticker mode")
    parser.add_argument("--output", default=None, help="Output PDF path for single-ticker mode")
    parser.add_argument("--wacc", type=float, default=None, help="Override WACC for DCF (e.g., 0.10)")
    parser.add_argument("--fcf-growth", type=float, default=None, help="Override FCF growth rate for DCF (e.g., 0.08)")
    parser.add_argument("--terminal-growth", type=float, default=None, help="Terminal growth rate for DCF (default: 0.025)")
    parser.add_argument(
        "--export-excel",
        nargs="?",
        const="",
        default=None,
        metavar="PATH",
        help="Export Simon-style DCF Excel (default: reports/{TICKER}_Simon_Valuation.xlsx)",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    if args.weekly and not args.force and not is_monday():
        print("Weekly mode: today is not Monday. Skipping. Use --force to override.")
        return 0

    if args.batch_top20:
        print("=" * 60)
        mode = "refresh-stale" if args.refresh_stale and not args.force else "full"
        print(f"  Batch ({mode}): {args.pdf_count} PDFs + Top {args.summary_count} Summary")
        print("=" * 60)
        rows = run_batch(
            output_dir=args.output_dir,
            wacc=args.wacc,
            replace_old=args.replace_old,
            pdf_count=args.pdf_count,
            summary_count=args.summary_count,
            refresh_stale=args.refresh_stale,
            force=args.force,
            max_age_days=args.stale_days,
        )
        print(f"\nCompleted: summary ({len(rows)} rows) in {args.output_dir}/")
        return 0 if rows else 1

    if not args.ticker:
        print("Error: provide a ticker or use --batch-top20", file=sys.stderr)
        return 1

    ticker = args.ticker.upper().strip()
    print(f"Fetching data for {ticker} from Yahoo Finance...")
    try:
        data = fetch_company(ticker)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    print("Building financial summary...")
    summary = build_financial_summary(data)
    industry = build_industry_research(data, summary)
    competitors = build_competitor_analysis(data, summary, peer_tickers=args.peers)
    valuation = build_valuation_summary(data, summary, competitors, wacc=args.wacc)

    if args.fcf_growth is not None or args.terminal_growth is not None:
        valuation.dcf = run_dcf(
            data, summary,
            wacc=args.wacc,
            terminal_growth=args.terminal_growth,
            fcf_growth=args.fcf_growth,
        )

    if args.output:
        output_path = args.output
    else:
        os.makedirs(args.output_dir, exist_ok=True)
        output_path = os.path.join(args.output_dir, f"{ticker}_research.pdf")

    if os.path.exists(output_path):
        try:
            os.remove(output_path)
        except OSError:
            base, ext = os.path.splitext(output_path)
            output_path = f"{base}_{datetime.now().strftime('%H%M%S')}{ext}"

    print(f"Generating PDF report -> {output_path}")
    generate_pdf_report(summary, industry, competitors, valuation, data.history, output_path, data)

    os.makedirs(args.output_dir, exist_ok=True)
    state = load_data_state(args.output_dir)
    row = _report_row_from_analysis(ticker, summary, valuation, output_path)
    record_ticker_state(state, ticker, data, row)
    save_data_state(args.output_dir, state)

    if args.export_excel is not None:
        excel_path = args.export_excel if args.export_excel else os.path.join(
            args.output_dir, f"{ticker}_Simon_Valuation.xlsx"
        )
        print(f"Exporting Simon DCF Excel -> {excel_path}")
        export_simon_valuation_excel(ticker, data, summary, valuation, competitors, excel_path)

    print("\n" + "=" * 50)
    print(f"  {summary.company_name} ({ticker})")
    print(f"  Price: ${summary.current_price:.2f}" if summary.current_price else "  Price: N/A")
    print(f"  Rating: {recommendation_short(valuation)}")
    print(f"  Recommendation: {valuation.recommendation}")
    if valuation.target_price:
        print(f"  Target Price: ${valuation.target_price:.2f}")
    print(f"  PDF saved: {output_path}")
    print("=" * 50)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
